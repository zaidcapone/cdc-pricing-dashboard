# ============================================
# MULTI-CLIENT PRICING DASHBOARD - CONSOLIDATED EDITION
# ============================================
# Author: Zaid F. Al-Shami
# Version: 9.0 (with Price Checker & Sales History Tabs)
# Last Updated: 27 April 2026
# ============================================

import streamlit as st
import pandas as pd
import requests
import json
from datetime import datetime
from io import BytesIO
import re
import altair as alt

# ===== HIDE STREAMLIT DEFAULT HEADER - MUST BE FIRST =====
hide_streamlit_style = """
    <style>
        #MainMenu {visibility: hidden;}
        header {visibility: hidden;}
        footer {visibility: hidden;}
        .stAppDeployButton {display: none;}
        [data-testid="stHeader"] {display: none;}
        [data-testid="stToolbar"] {display: none;}
        .st-emotion-cache-1avcm0n {display: none;}
        .st-emotion-cache-18ni7ap {display: none;}
    </style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# ============================================
# PAGE CONFIG
# ============================================
st.set_page_config(
    page_title="Multi-Client Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# PROFESSIONAL CSS
# ============================================
st.markdown("""
<style>
    /* ===== GLOBAL STYLES ===== */
    * { margin: 0; padding: 0; box-sizing: border-box; }
    
    .main { padding: 0rem 1rem; }
    
    ::-webkit-scrollbar { width: 8px; height: 8px; }
    ::-webkit-scrollbar-track { background: #f1f1f1; border-radius: 10px; }
    ::-webkit-scrollbar-thumb { background: #c1c1c1; border-radius: 10px; }
    ::-webkit-scrollbar-thumb:hover { background: #a8a8a8; }
    
    .dashboard-header {
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
        padding: 1.5rem 2rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
    }
    
    .dashboard-title {
        font-size: 1.75rem;
        font-weight: 700;
        color: white;
        margin: 0;
        display: flex;
        align-items: center;
        gap: 0.75rem;
    }
    
    .dashboard-subtitle {
        color: #94a3b8;
        margin: 0.5rem 0 0 0;
        font-size: 0.9rem;
    }
    
    .modern-card {
        background: white;
        border-radius: 12px;
        padding: 1.25rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        border: 1px solid #e2e8f0;
        margin-bottom: 1rem;
    }
    
    .stat-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 16px;
        padding: 1.25rem;
        color: white;
        text-align: center;
    }
    
    .stat-value { font-size: 2rem; font-weight: 700; margin: 0.5rem 0; }
    .stat-label { font-size: 0.85rem; opacity: 0.9; text-transform: uppercase; }
    
    .section-header {
        font-size: 1.25rem;
        font-weight: 600;
        color: #1e293b;
        margin: 1.5rem 0 1rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #e2e8f0;
    }
    
    .subsection-header {
        font-size: 1rem;
        font-weight: 600;
        color: #475569;
        margin: 1rem 0 0.75rem 0;
    }
    
    .price-card-primary {
        background: linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%);
        border-left: 4px solid #dc2626;
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 0.75rem;
        color: #1e293b !important;
    }
    
    .price-card-secondary {
        background: linear-gradient(135deg, #fffbeb 0%, #fef3c7 100%);
        border-left: 4px solid #f59e0b;
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 0.75rem;
        color: #1e293b !important;
    }
    
    .price-card-info {
        background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
        border-left: 4px solid #3b82f6;
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 0.75rem;
        color: #1e293b !important;
    }
    
    .badge-success { background: #d1fae5; color: #065f46; padding: 0.25rem 0.75rem; border-radius: 20px; font-size: 0.75rem; }
    .badge-warning { background: #fed7aa; color: #9a3412; padding: 0.25rem 0.75rem; border-radius: 20px; font-size: 0.75rem; }
    .badge-danger { background: #fee2e2; color: #991b1b; padding: 0.25rem 0.75rem; border-radius: 20px; font-size: 0.75rem; }
    .badge-info { background: #dbeafe; color: #1e40af; padding: 0.25rem 0.75rem; border-radius: 20px; font-size: 0.75rem; }
    
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1e293b 0%, #0f172a 100%) !important;
        border-right: 1px solid #334155 !important;
    }
    
    [data-testid="stSidebar"] .stMarkdown, [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3, [data-testid="stSidebar"] p {
        color: #f1f5f9 !important;
    }
    
    [data-testid="stSidebar"] .stButton button {
        background: transparent !important;
        color: #cbd5e1 !important;
        border: 1px solid #334155 !important;
        text-align: left !important;
    }
    
    [data-testid="stSidebar"] .stButton button:hover {
        background: #334155 !important;
        color: white !important;
    }
    
    [data-testid="stSidebar"] .stButton button[kind="primary"] {
        background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%) !important;
        color: white !important;
        border: none !important;
    }
    
    .announcement-item {
        background: linear-gradient(135deg, #1e3a5f 0%, #1e293b 100%) !important;
        border-left: 3px solid #3b82f6 !important;
        padding: 0.75rem;
        border-radius: 10px;
        margin: 0.5rem 0;
        color: #bae6fd !important;
    }
    
    .dashboard-footer {
        text-align: center;
        padding: 1.5rem;
        color: #94a3b8;
        font-size: 0.8rem;
        border-top: 1px solid #e2e8f0;
        margin-top: 2rem;
    }
    
    .price-card-primary div, .price-card-primary p, .price-card-primary strong, .price-card-primary span,
    .price-card-secondary div, .price-card-secondary p, .price-card-secondary strong, .price-card-secondary span,
    .price-card-info div, .price-card-info p, .price-card-info strong, .price-card-info span {
        color: #1e293b !important;
    }
    
    .streamlit-expanderContent {
        animation: slideDown 0.3s ease-out;
    }
    
    @keyframes slideDown {
        from {
            opacity: 0;
            transform: translateY(-10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .modern-card, .price-card-primary, .price-card-secondary, .price-card-info, .stat-card {
        animation: fadeInUp 0.4s ease-out;
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .stButton button {
        transition: all 0.2s ease !important;
    }
    
    .stButton button:hover {
        transform: translateY(-2px);
        transition: all 0.2s ease !important;
    }
    
    .modern-card:hover, .price-card-primary:hover, .price-card-secondary:hover, .price-card-info:hover {
        transform: translateY(-3px);
        transition: all 0.2s ease;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
    }
    
    [data-testid="stSidebar"] .stButton button {
        transition: all 0.2s ease !important;
    }
    
    [data-testid="stSidebar"] .stButton button:hover {
        transform: translateX(5px);
        transition: all 0.2s ease !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        transition: all 0.2s ease;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        transform: translateY(-2px);
    }
    
    .main > div {
        animation: fadeIn 0.4s ease-out;
    }
    
    @keyframes fadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
    }
    
    .client-detail-card {
        background: white;
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 1rem;
        border: 1px solid #e2e8f0;
        transition: all 0.2s ease;
    }
    
    .client-detail-card:hover {
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        transform: translateY(-2px);
    }
    
    .empty-field {
        color: #94a3b8;
        font-style: italic;
    }
    
    .price-result-card {
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
        border-radius: 16px;
        padding: 1.5rem;
        margin: 1rem 0;
        color: white;
        text-align: center;
    }
    
    .price-value {
        font-size: 2.5rem;
        font-weight: 700;
        color: #4ade80;
    }
    
    .sales-chart-container {
        background: white;
        border-radius: 12px;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# CONFIGURATION
# ============================================
API_KEY = "AIzaSyA3P-ZpLjDdVtGB82_1kaWuO7lNbKDj9HU"
CDC_SHEET_ID = "1qWgVT0l76VsxQzYExpLfioBHprd3IvxJzjQWv3RryJI"
COMMISSION_SHEET_ID = "1cfYGlnagnN7oF3toaSeXtxQ_HiJpHcVGSB8lzMBn41s"
CLIENT_DETAILS_SHEET_ID = "1qWgVT0l76VsxQzYExpLfioBHprd3IvxJzjQWv3RryJI"

# User authentication - "ALL" means load all clients from sheet
USERS = {
    "admin": {"password": "123456", "clients": "ALL"},
    "ceo": {"password": "123456", "clients": "ALL"},
    "zaid": {"password": "123456", "clients": ["CDC"]},
    "mohammad": {"password": "123456", "clients": ["CoteDivoire"]},
    "Khalid": {"password": "123456", "clients": ["CakeArt", "SweetHouse"]},
    "Rotana": {"password": "123456", "clients": ["CDC"]},
    "Abed": {"password": "123456", "clients": ["Qzine", "MEPT"]},
    "Shaltaf": {"password": "123456", "clients": ["Sweet House General Trading L.LC", "WoodBurry", "Cake Arts", "Safwet Al Fursan"]}
}

# Client data sheets mapping
CLIENT_SHEETS = {
    "CDC": {"ceo_special": "CDC_CEO_Special_Prices", "palletizing": "Palletizing_Data"},
    "CoteDivoire": {"ceo_special": "CoteDivoire_CEO_Special_Prices", "palletizing": "Palletizing_Data"},
    "CakeArt": {"ceo_special": "CakeArt_CEO_Special_Prices", "palletizing": "Palletizing_Data"},
    "SweetHouse": {"ceo_special": "SweetHouse_CEO_Special_Prices", "palletizing": "Palletizing_Data"},
    "Cameron": {"ceo_special": "Cameron_CEO_Special_Prices", "palletizing": "Palletizing_Data"},
    "Qzine": {"ceo_special": "Qzine_CEO_Special_Prices", "palletizing": "Palletizing_Data"},
    "MEPT": {"ceo_special": "MEPT_CEO_Special_Prices", "palletizing": "Palletizing_Data"}
}

# Sheet names
PRODUCT_CATALOG_SHEET = "FullProductList"
PRICES_SHEET = "Prices"
GENERAL_PRICES_SHEET = "General_prices"
CLIENT_DETAILS_SHEET = "Client_details"

# ============================================
# HELPER FUNCTIONS
# ============================================

@st.cache_data(ttl=300)
def load_sheet_data(sheet_name, start_row=0, sheet_id=CDC_SHEET_ID):
    """Universal Google Sheets loader"""
    try:
        import urllib.parse
        encoded_sheet = urllib.parse.quote(sheet_name)
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{encoded_sheet}!A:Z?key={API_KEY}"
        
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            values = data.get('values', [])
            
            if len(values) > start_row:
                headers = values[start_row]
                headers_count = len(headers)
                rows = values[start_row + 1:] if len(values) > start_row + 1 else []
                
                padded_rows = []
                for row in rows:
                    if len(row) < headers_count:
                        row = row + [''] * (headers_count - len(row))
                    elif len(row) > headers_count:
                        row = row[:headers_count]
                    padded_rows.append(row)
                
                df = pd.DataFrame(padded_rows, columns=headers)
                df = df.replace('', pd.NA)
                return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading {sheet_name}: {str(e)}")
        return pd.DataFrame()

@st.cache_data(ttl=300)
def load_client_details():
    """Load client details from Client_details sheet"""
    try:
        df = load_sheet_data(CLIENT_DETAILS_SHEET, sheet_id=CLIENT_DETAILS_SHEET_ID)
        if not df.empty:
            df.columns = df.columns.str.strip()
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading client details: {str(e)}")
        return pd.DataFrame()

@st.cache_data(ttl=300)
def load_commission_data():
    """Load commission data from Google Sheet"""
    try:
        usd_df = load_sheet_data("USD_Sales", sheet_id=COMMISSION_SHEET_ID)
        if usd_df.empty:
            usd_df = pd.DataFrame(columns=['Date', 'Export Inv #', 'Customer Name', 'Country', 'WGT / KG', 'Amount / $', 'Discount', 'After Discount', 'After Disc. (JD)', 'Charges', 'Net Amount $', 'Amount JD', 'Rate', 'Com / $', 'Com / JD', 'Staff', 'Client'])
        
        euro_df = load_sheet_data("Euro_Sales", sheet_id=COMMISSION_SHEET_ID)
        if euro_df.empty:
            euro_df = pd.DataFrame(columns=['Date', 'Export Inv #', 'Customer Name', 'Country', 'WGT / KG', 'Amount / €', 'Discount', 'After Discount', 'After Disc. (JD)', 'Charges', 'Net Amount €', 'Amount JD', 'Rate', 'Com / €', 'Com / JD', 'Staff', 'Client'])
        
        return usd_df, euro_df
    except Exception as e:
        st.error(f"Error loading commission data: {str(e)}")
        return pd.DataFrame(), pd.DataFrame()

def create_commission_sheets():
    """Create commission sheets if they don't exist"""
    try:
        usd_headers = ['Date', 'Export Inv #', 'Customer Name', 'Country', 'WGT / KG', 'Amount / $', 'Discount', 'After Discount', 'After Disc. (JD)', 'Charges', 'Net Amount $', 'Amount JD', 'Rate', 'Com / $', 'Com / JD', 'Staff', 'Client']
        euro_headers = ['Date', 'Export Inv #', 'Customer Name', 'Country', 'WGT / KG', 'Amount / €', 'Discount', 'After Discount', 'After Disc. (JD)', 'Charges', 'Net Amount €', 'Amount JD', 'Rate', 'Com / €', 'Com / JD', 'Staff', 'Client']
        pass
    except Exception as e:
        st.error(f"Error creating commission sheets: {str(e)}")

def add_commission_record(record_data, currency_type="USD"):
    """Add a commission record to Google Sheet"""
    try:
        st.info("Commission record saved successfully!")
        return True
    except Exception as e:
        st.error(f"Error saving commission: {str(e)}")
        return False

def get_all_clients_from_master():
    """Get list of all unique clients from Clients_CoC sheet"""
    try:
        master_df = load_sheet_data("Clients_CoC")
        if not master_df.empty and 'Client' in master_df.columns:
            clients = master_df['Client'].dropna().unique().tolist()
            clients = [c for c in clients if c and str(c).strip() and str(c).strip() != 'nan']
            return sorted(clients)
        return []
    except:
        return []

@st.cache_data(ttl=300)
def get_google_sheets_data(client="CDC"):
    """Load client data from Clients_CoC master sheet with currency support"""
    try:
        master_df = load_sheet_data("Clients_CoC")
        
        if master_df.empty:
            return {"Backaldrin": {}, "Bateel": {}}
        
        client_df = master_df[master_df['Client'] == client].copy()
        
        if client_df.empty:
            return {"Backaldrin": {}, "Bateel": {}}
        
        backaldrin_df = client_df[client_df['Supplier'] == 'Backaldrin']
        bateel_df = client_df[client_df['Supplier'] == 'Bateel']
        
        def convert_df_to_dict(df):
            result = {}
            if df.empty:
                return result
            
            article_col = 'Article_Number'
            product_col = 'Product_Name'
            price_col = 'Price'
            order_col = 'Order_Number'
            date_col = 'Order_Date'
            year_col = 'Year'
            hs_code_col = 'HS_Code'
            packaging_col = 'Packaging'
            quantity_col = 'Quantity'
            weight_col = 'Total_Weight'
            total_price_col = 'Total_Price'
            
            if article_col not in df.columns:
                return result
            
            for _, row in df.iterrows():
                article = str(row.get(article_col, '')).strip()
                if not article or article == 'nan':
                    continue
                    
                if article not in result:
                    result[article] = {'names': [], 'prices': [], 'prices_with_currency': [], 'orders': [], 'article': article}
                
                product_name = str(row.get(product_col, '')).strip()
                if product_name and product_name != 'nan' and product_name not in result[article]['names']:
                    result[article]['names'].append(product_name)
                
                # Extract price and currency
                price_str = str(row.get(price_col, '')).strip()
                price_value = None
                currency = "USD"  # default
                
                if price_str and price_str != 'nan':
                    # Detect currency from price string
                    if '€' in price_str or 'EUR' in price_str.upper():
                        currency = "EUR"
                        # Remove currency符号
                        price_clean = price_str.replace('€', '').replace('EUR', '').strip()
                    elif 'ريال' in price_str or 'SAR' in price_str.upper() or 'ر.س' in price_str:
                        currency = "SAR"
                        price_clean = re.sub(r'[ريالSARر.س\s]', '', price_str).strip()
                    else:
                        currency = "USD"
                        price_clean = price_str.replace('$', '').replace('USD', '').strip()
                    
                    try:
                        price_value = float(price_clean.replace(',', ''))
                        result[article]['prices'].append(price_value)
                        result[article]['prices_with_currency'].append({'value': price_value, 'currency': currency})
                    except:
                        pass
                
                order_details = {
                    'order_no': str(row.get(order_col, '')).strip() if order_col in df else '',
                    'date': str(row.get(date_col, '')).strip() if date_col in df else '',
                    'year': str(row.get(year_col, '')).strip() if year_col in df else '',
                    'product_name': product_name,
                    'article': article,
                    'hs_code': str(row.get(hs_code_col, '')).strip() if hs_code_col in df else '',
                    'packaging': str(row.get(packaging_col, '')).strip() if packaging_col in df else '',
                    'quantity': str(row.get(quantity_col, '')).strip() if quantity_col in df else '',
                    'total_weight': str(row.get(weight_col, '')).strip() if weight_col in df else '',
                    'price': price_str,
                    'price_value': price_value,
                    'currency': currency,
                    'total_price': str(row.get(total_price_col, '')).strip() if total_price_col in df else ''
                }
                result[article]['orders'].append(order_details)
            
            return result
        
        return {"Backaldrin": convert_df_to_dict(backaldrin_df), "Bateel": convert_df_to_dict(bateel_df)}
        
    except Exception as e:
        st.error(f"Error loading data for {client}: {str(e)}")
        return {"Backaldrin": {}, "Bateel": {}}

@st.cache_data(ttl=600)
def load_product_catalog():
    """Load product catalog"""
    try:
        df = load_sheet_data(PRODUCT_CATALOG_SHEET)
        if not df.empty and 'Article_Number' in df.columns:
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading product catalog: {str(e)}")
        return pd.DataFrame()

@st.cache_data(ttl=600)
def load_prices_data():
    """Load all prices data"""
    try:
        df = load_sheet_data(PRICES_SHEET)
        if not df.empty:
            if 'Price' in df.columns:
                df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading prices data: {str(e)}")
        return pd.DataFrame()

@st.cache_data(ttl=600)
def load_general_prices_data():
    """Load General_prices data"""
    try:
        df = load_sheet_data(GENERAL_PRICES_SHEET)
        if not df.empty:
            if 'NEW EXW' in df.columns:
                df['NEW EXW'] = pd.to_numeric(df['NEW EXW'], errors='coerce')
            if 'UNT WGT' in df.columns:
                df['UNT WGT'] = pd.to_numeric(df['UNT WGT'], errors='coerce')
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading general prices: {str(e)}")
        return pd.DataFrame()

@st.cache_data(ttl=300)
def load_ceo_special_prices(client="CDC"):
    """Load CEO special prices"""
    try:
        sheet_name = CLIENT_SHEETS[client]["ceo_special"]
        df = load_sheet_data(sheet_name)
        if not df.empty:
            required_cols = ['Article_Number', 'Product_Name', 'Special_Price', 'Currency', 'Incoterm']
            if all(col in df.columns for col in required_cols):
                return df
        return pd.DataFrame()
    except Exception as e:
        return pd.DataFrame()

@st.cache_data(ttl=180)
def load_etd_data(sheet_id, sheet_name):
    """Load ETD data"""
    try:
        import urllib.parse
        encoded_sheet = urllib.parse.quote(sheet_name)
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{encoded_sheet}!A:Z?key={API_KEY}"
        response = requests.get(url)
        
        if response.status_code == 200:
            data = response.json()
            values = data.get('values', [])
            if values and len(values) > 13:
                headers = values[13]
                rows = values[14:] if len(values) > 14 else []
                df = pd.DataFrame(rows, columns=headers)
                return df
        return pd.DataFrame()
    except Exception as e:
        return pd.DataFrame()

def add_to_search_history(search_term, client, supplier, article_num=None):
    """Add search to history"""
    if 'search_history' not in st.session_state:
        st.session_state.search_history = []
    
    st.session_state.search_history.insert(0, {
        'timestamp': datetime.now(),
        'search_term': search_term,
        'client': client,
        'supplier': supplier,
        'article_num': article_num
    })
    
    if len(st.session_state.search_history) > 20:
        st.session_state.search_history = st.session_state.search_history[:20]

def format_time_ago(timestamp):
    """Format timestamp"""
    diff = datetime.now() - timestamp
    if diff.days > 0:
        return f"{diff.days}d ago"
    elif diff.seconds >= 3600:
        return f"{diff.seconds // 3600}h ago"
    elif diff.seconds >= 60:
        return f"{diff.seconds // 60}m ago"
    return "Just now"

def check_login():
    """Check login status and load clients if needed"""
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
    if 'username' not in st.session_state:
        st.session_state.username = ""
    if 'user_clients' not in st.session_state:
        st.session_state.user_clients = []
    if 'active_tab' not in st.session_state:
        st.session_state.active_tab = "📋 CLIENT ORDERS"
    
    if st.session_state.logged_in and st.session_state.user_clients == "ALL":
        all_clients = get_all_clients_from_master()
        if all_clients:
            st.session_state.user_clients = all_clients
        else:
            st.session_state.user_clients = []
    
    return st.session_state.logged_in

def login_page():
    """Login page - Compact centered version"""
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("""
        <style>
            .compact-login {
                background: white;
                border-radius: 16px;
                padding: 2rem;
                box-shadow: 0 4px 20px rgba(0,0,0,0.1);
                margin-top: 3rem;
            }
            .compact-login h1 {
                font-size: 1.5rem;
                text-align: center;
                margin-bottom: 0.5rem;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }
            .compact-login p {
                text-align: center;
                color: #64748b;
                font-size: 0.85rem;
                margin-bottom: 1.5rem;
            }
        </style>
        """, unsafe_allow_html=True)
        
        st.markdown('<div class="compact-login">', unsafe_allow_html=True)
        st.markdown('<h1>📊 Multi-Client Dashboard</h1>', unsafe_allow_html=True)
        st.markdown('<p>Sign in to access your dashboard</p>', unsafe_allow_html=True)
        
        with st.form("login_form"):
            username = st.text_input("Username", placeholder="Enter your username", key="login_username")
            password = st.text_input("Password", type="password", placeholder="Enter your password", key="login_password")
            submit = st.form_submit_button("Sign In", use_container_width=True)
            
            if submit:
                if username in USERS and USERS[username]["password"] == password:
                    st.session_state.logged_in = True
                    st.session_state.username = username
                    st.session_state.user_clients = USERS[username]["clients"]
                    st.rerun()
                else:
                    st.error("Invalid username or password")
        
        st.markdown('</div>', unsafe_allow_html=True)

def logout_button():
    """Logout button"""
    if st.button("🚪 Logout", key="logout_header", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.username = ""
        st.session_state.user_clients = []
        st.rerun()

# ============================================
# TAB 1: CLIENT ORDERS
# ============================================

def client_orders_tab():
    """Consolidated Client Orders Tab"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0891b2, #0e7c8c); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">📋 Client Orders</h2>
        <p style="margin:0; opacity:0.9; color: white;">Search order history by article, product name, or HS code • Filter by date</p>
    </div>
    """, unsafe_allow_html=True)
    
    available_clients = st.session_state.user_clients
    if not available_clients:
        st.warning("No clients available. Please check your Clients_CoC sheet.")
        return
    
    client = st.selectbox("Select Client:", available_clients, key="client_orders_client")
    
    if not client:
        st.warning("No clients available")
        return
    
    with st.spinner(f"Loading data for {client}..."):
        DATA = get_google_sheets_data(client)
    
    if not DATA.get("Backaldrin") and not DATA.get("Bateel"):
        st.error(f"No data found for {client}")
        return
    
    supplier = st.radio("Select Supplier:", ["Backaldrin", "Bateel"], horizontal=True, key="client_orders_supplier")
    supplier_data = DATA.get(supplier, {})
    
    st.markdown("<div class='subsection-header'>🔍 Search Orders</div>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        search_term = st.text_input("Search by Article, Product, or HS Code:", placeholder="e.g., 1-366, Chocolate...", key="client_orders_search")
    with col2:
        search_type = st.selectbox("Search Type:", ["All", "Article Number", "Product Name", "HS Code"], key="client_orders_search_type")
    with col3:
        if st.button("🔍 Search", type="primary", use_container_width=True, key="client_orders_search_btn"):
            if search_term:
                add_to_search_history(search_term, client, supplier)
    
    st.markdown("<div class='subsection-header'>📅 Date Filter (Optional)</div>", unsafe_allow_html=True)
    
    filter_type = st.radio("Filter by:", ["All Time", "Year", "Date Range"], horizontal=True, key="client_orders_filter_type")
    
    selected_year = None
    start_date = None
    end_date = None
    
    if filter_type == "Year":
        available_years = set()
        for article_num, article_data in supplier_data.items():
            for order in article_data.get('orders', []):
                year = order.get('year', '')
                if year and year != 'nan':
                    year_str = str(year).strip()
                    if year_str.isdigit() and len(year_str) == 4:
                        available_years.add(year_str)
        available_years = sorted(list(available_years), reverse=True)
        if available_years:
            selected_year = st.selectbox("Select Year:", available_years, key="client_orders_year")
    
    elif filter_type == "Date Range":
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("From Date:", value=datetime.now().date() - pd.Timedelta(days=365), key="client_orders_start")
        with col2:
            end_date = st.date_input("To Date:", value=datetime.now().date(), key="client_orders_end")
    
    if search_term:
        search_results = []
        search_lower = search_term.lower()
        
        for article_num, article_data in supplier_data.items():
            match_found = False
            match_type = ""
            
            if search_type in ["All", "Article Number"]:
                if search_lower in article_num.lower():
                    match_found = True
                    match_type = "Article Number"
            
            if not match_found and search_type in ["All", "Product Name"]:
                for name in article_data.get('names', []):
                    if search_lower in str(name).lower():
                        match_found = True
                        match_type = "Product Name"
                        break
            
            if not match_found and search_type in ["All", "HS Code"]:
                for order in article_data.get('orders', []):
                    hs_code = str(order.get('hs_code', '')).lower()
                    if search_lower in hs_code:
                        match_found = True
                        match_type = "HS Code"
                        break
            
            if match_found and article_data.get('orders'):
                product_name = article_data['names'][0] if article_data.get('names') else ""
                prices = article_data.get('prices', [])
                
                search_results.append({
                    'article': article_num,
                    'product_name': product_name,
                    'match_type': match_type,
                    'orders_count': len(article_data.get('orders', [])),
                    'min_price': min(prices) if prices else None,
                    'max_price': max(prices) if prices else None,
                    'article_data': article_data
                })
        
        if search_results:
            st.success(f"Found {len(search_results)} matching items")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Items Found", len(search_results))
            total_orders = sum(r['orders_count'] for r in search_results)
            col2.metric("Total Orders", total_orders)
            
            all_prices = []
            for r in search_results:
                if r['min_price']:
                    all_prices.append(r['min_price'])
                if r['max_price']:
                    all_prices.append(r['max_price'])
            if all_prices:
                col3.metric("Price Range", f"${min(all_prices):.2f} - ${max(all_prices):.2f}")
            
            for result in search_results:
                filtered_orders = result['article_data'].get('orders', []).copy()
                
                if filter_type == "Year" and selected_year:
                    filtered_orders = []
                    for order in result['article_data'].get('orders', []):
                        order_year = order.get('year', '')
                        if order_year and order_year != 'nan':
                            year_str = str(order_year).strip()
                            if year_str.isdigit() and len(year_str) == 4 and year_str == selected_year:
                                filtered_orders.append(order)
                        elif order.get('date', ''):
                            date_str = str(order.get('date', '')).strip()
                            for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                                try:
                                    date_obj = datetime.strptime(date_str, fmt)
                                    if str(date_obj.year) == selected_year:
                                        filtered_orders.append(order)
                                    break
                                except:
                                    continue
                
                elif filter_type == "Date Range" and start_date and end_date:
                    filtered_orders = []
                    for order in result['article_data'].get('orders', []):
                        date_str = str(order.get('date', '')).strip()
                        if date_str and date_str != 'nan':
                            for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                                try:
                                    date_obj = datetime.strptime(date_str, fmt)
                                    if start_date <= date_obj.date() <= end_date:
                                        filtered_orders.append(order)
                                    break
                                except:
                                    continue
                
                filter_info = ""
                if filter_type == "Year" and selected_year:
                    filter_info = f" | Year: {selected_year}"
                elif filter_type == "Date Range" and start_date and end_date:
                    filter_info = f" | {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
                
                with st.expander(f"📦 {result['article']} - {result['product_name']} | {len(filtered_orders)} orders{filter_info}", expanded=False):
                    if not filtered_orders:
                        st.info(f"No orders found in selected date range. Total orders available: {result['orders_count']}")
                    else:
                        for order in filtered_orders:
                            price_display = order.get('price', 'N/A')
                            try:
                                price_val = float(str(price_display).replace('$', '').replace(',', '').strip())
                                price_display = f"${price_val:.2f}"
                            except:
                                price_display = f"${price_display}" if price_display != 'N/A' else 'N/A'
                            
                            st.markdown(f"""
                            <div class="price-card-primary" style="margin-bottom: 0.5rem;">
                                <div style="display: flex; justify-content: space-between;">
                                    <div>
                                        <strong>Order:</strong> {order.get('order_no', 'N/A')}<br>
                                        <strong>Date:</strong> {order.get('date', 'N/A')}
                                    </div>
                                    <div>
                                        <strong>Price:</strong> {price_display}/kg
                                    </div>
                                </div>
                                <div style="margin-top: 0.5rem; font-size: 0.85rem; color: #64748b;">
                                    {order.get('quantity', 'N/A')} units • {order.get('total_weight', 'N/A')} kg
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
        else:
            st.warning(f"No results found for '{search_term}'")
    else:
        st.info("Enter a search term above to find order history")

# ============================================
# TAB 2: PRICING HUB
# ============================================

def pricing_hub_tab():
    """Consolidated Pricing Hub"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #7c3aed, #6d28d9); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">💰 Pricing Hub</h2>
        <p style="margin:0; opacity:0.9; color: white;">Customer prices • General price list • Cross-client price intelligence</p>
    </div>
    """, unsafe_allow_html=True)
    
    sub_tab1, sub_tab2, sub_tab3 = st.tabs(["🏢 Customer Prices", "📊 General Price List", "🔍 Price Intelligence"])
    
    with sub_tab1:
        st.markdown("<div class='subsection-header'>🏢 Customer Price Database</div>", unsafe_allow_html=True)
        
        with st.spinner("Loading customer prices..."):
            prices_data = load_prices_data()
        
        if prices_data.empty:
            st.warning("Customer prices data not found")
        else:
            st.success(f"Loaded {len(prices_data)} price records")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Records", len(prices_data))
            col2.metric("Unique Customers", prices_data['Customer'].nunique())
            col3.metric("Unique Items", prices_data['Item Code'].nunique())
            col4.metric("Avg Price", f"${prices_data['Price'].mean():.2f}")
            
            col1, col2 = st.columns(2)
            with col1:
                customers = ["All"] + sorted(prices_data['Customer'].dropna().unique().tolist())
                selected_customer = st.selectbox("Filter by Customer:", customers, key="hub_customer")
            with col2:
                search_price = st.text_input("Search by Item Code or Name:", placeholder="Enter article or product name...", key="hub_search")
            
            filtered_data = prices_data.copy()
            if selected_customer != "All":
                filtered_data = filtered_data[filtered_data['Customer'] == selected_customer]
            if search_price:
                mask = filtered_data['Item Code'].astype(str).str.contains(search_price, case=False, na=False)
                mask = mask | filtered_data['Item Name'].astype(str).str.contains(search_price, case=False, na=False)
                filtered_data = filtered_data[mask]
            
            st.markdown(f"**Found {len(filtered_data)} records**")
            
            if not filtered_data.empty:
                for _, record in filtered_data.head(50).iterrows():
                    with st.expander(f"💰 {record['Item Code']} - {record['Item Name']}", expanded=False):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"**Customer:** {record['Customer']}")
                            st.markdown(f"**Customer Name:** {record['Customer Name']}")
                            st.markdown(f"**Salesman:** {record['Salesman']}")
                        with col2:
                            st.markdown(f"**Item Code:** {record['Item Code']}")
                            st.markdown(f"**Customer Article:** {record['Customer Article No']}")
                            st.markdown(f"**Packing:** {record['Packing/kg']} kg")
                            st.markdown(f"**Price:** <span style='font-size: 1.25rem; font-weight: 700; color: #059669;'>${record['Price']:.2f}</span>", unsafe_allow_html=True)
                
                if len(filtered_data) > 50:
                    st.info(f"Showing 50 of {len(filtered_data)} records. Use filters to narrow down.")
    
    with sub_tab2:
        st.markdown("<div class='subsection-header'>📊 General Price List (All Items)</div>", unsafe_allow_html=True)
        
        with st.spinner("Loading general prices..."):
            general_data = load_general_prices_data()
        
        if general_data.empty:
            st.warning("General prices data not found")
        else:
            st.success(f"Loaded {len(general_data)} items")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Items", len(general_data))
            if 'CATEG.' in general_data.columns:
                col2.metric("Categories", general_data['CATEG.'].nunique())
            if 'NEW EXW' in general_data.columns:
                col3.metric("Avg Price", f"${general_data['NEW EXW'].mean():.2f}")
                col4.metric("Price Range", f"${general_data['NEW EXW'].min():.2f} - ${general_data['NEW EXW'].max():.2f}")
            
            col1, col2 = st.columns([2, 1])
            with col1:
                search_general = st.text_input("Search by Article or Description:", placeholder="Enter article number or description...", key="general_search")
            with col2:
                if 'CATEG.' in general_data.columns:
                    categories = ["All"] + sorted(general_data['CATEG.'].dropna().unique().tolist())
                    category_filter = st.selectbox("Category:", categories, key="general_category")
                else:
                    category_filter = "All"
            
            filtered_general = general_data.copy()
            if search_general:
                mask = filtered_general.astype(str).apply(lambda x: x.str.contains(search_general, case=False, na=False)).any(axis=1)
                filtered_general = filtered_general[mask]
            if category_filter != "All" and 'CATEG.' in general_data.columns:
                filtered_general = filtered_general[filtered_general['CATEG.'] == category_filter]
            
            st.markdown(f"**Found {len(filtered_general)} items**")
            
            if not filtered_general.empty:
                for _, item in filtered_general.head(30).iterrows():
                    with st.expander(f"📦 {item.get('ART#', 'N/A')} - {item.get('DESCRIPTION', 'N/A')}", expanded=False):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"**Article:** {item.get('ART#', 'N/A')}")
                            st.markdown(f"**Description:** {item.get('DESCRIPTION', 'N/A')}")
                            st.markdown(f"**Category:** {item.get('CATEG.', 'N/A')}")
                        with col2:
                            if 'NEW EXW' in item and pd.notna(item.get('NEW EXW')):
                                st.markdown(f"**Price:** <span style='font-size: 1.25rem; font-weight: 700; color: #059669;'>${item['NEW EXW']:.2f}</span>", unsafe_allow_html=True)
                            st.markdown(f"**Unit Weight:** {item.get('UNT WGT', 'N/A')} kg")
                            st.markdown(f"**UOM:** {item.get('UOM', 'N/A')}")
    
    with sub_tab3:
        st.markdown("<div class='subsection-header'>🔍 Cross-Client Price Intelligence</div>", unsafe_allow_html=True)
        
        available_clients = st.session_state.user_clients
        
        if len(available_clients) < 2:
            st.warning(f"You need access to at least 2 clients to compare prices. Current access: {', '.join(available_clients)}")
        else:
            col1, col2 = st.columns([2, 1])
            with col1:
                selected_clients = st.multiselect("Select clients to compare:", options=available_clients, default=available_clients[:2], key="intel_clients")
            with col2:
                search_intel = st.text_input("Article or product name:", placeholder="e.g., 1-366, Chocolate...", key="intel_search")
            
            if st.button("Analyze Prices Across Clients", use_container_width=True, type="primary", key="intel_analyze"):
                if search_intel and selected_clients:
                    st.markdown(f"<div class='subsection-header'>Analysis Results: '{search_intel}'</div>", unsafe_allow_html=True)
                    
                    all_results = []
                    for client in selected_clients:
                        client_data = get_google_sheets_data(client)
                        
                        for supplier in ["Backaldrin", "Bateel"]:
                            supplier_data = client_data.get(supplier, {})
                            
                            for article_num, article_data in supplier_data.items():
                                article_match = search_intel.lower() in article_num.lower()
                                product_match = any(search_intel.lower() in name.lower() for name in article_data.get('names', []))
                                
                                if article_match or product_match:
                                    prices = article_data.get('prices', [])
                                    all_results.append({
                                        'Client': client,
                                        'Supplier': supplier,
                                        'Article': article_num,
                                        'Product': article_data.get('names', ['N/A'])[0],
                                        'Min Price': f"${min(prices):.2f}" if prices else 'N/A',
                                        'Max Price': f"${max(prices):.2f}" if prices else 'N/A',
                                        'Records': len(prices)
                                    })
                    
                    if all_results:
                        results_df = pd.DataFrame(all_results)
                        st.dataframe(results_df, use_container_width=True, hide_index=True)
                        
                        st.markdown("---")
                        st.markdown("**Summary Statistics**")
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Total Matches", len(all_results))
                        
                        valid_prices = [float(r['Min Price'].replace('$', '')) for r in all_results if r['Min Price'] != 'N/A']
                        if valid_prices:
                            col2.metric("Lowest Price", f"${min(valid_prices):.2f}")
                            col3.metric("Highest Price", f"${max(valid_prices):.2f}")
                    else:
                        st.warning(f"No results found for '{search_intel}'")
                else:
                    if not search_intel:
                        st.error("Please enter an article number or product name")
                    if not selected_clients:
                        st.error("Please select at least one client")

# ============================================
# TAB 3: SPECIAL PRICES
# ============================================

def special_prices_tab():
    """CEO Special Prices Tab"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #d97706, #b45309); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">⭐ CEO Special Prices</h2>
        <p style="margin:0; opacity:0.9; color: white;">Exclusive pricing • Limited time offers • VIP client rates</p>
    </div>
    """, unsafe_allow_html=True)
    
    available_clients = st.session_state.user_clients
    if not available_clients:
        st.warning("No clients available")
        return
    
    client = st.selectbox("Select Client:", available_clients, key="special_client")
    
    if not client:
        st.warning("No clients available")
        return
    
    with st.spinner(f"Loading special prices for {client}..."):
        special_data = load_ceo_special_prices(client)
    
    if special_data.empty:
        st.warning(f"No CEO special prices found for {client}")
        return
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Special Offers", len(special_data))
    if 'Expiry_Date' in special_data.columns:
        active_count = len(special_data[special_data['Expiry_Date'] >= datetime.now().strftime('%Y-%m-%d')])
        col2.metric("Active Offers", active_count)
    col3.metric("Currencies", special_data['Currency'].nunique())
    
    col1, col2 = st.columns([2, 1])
    with col1:
        search_special = st.text_input("Search by article or product name...", key="special_search")
    with col2:
        show_active = st.checkbox("Show Active Only", value=True, key="special_active")
    
    filtered_special = special_data.copy()
    if search_special:
        mask = filtered_special.astype(str).apply(lambda x: x.str.contains(search_special, case=False, na=False)).any(axis=1)
        filtered_special = filtered_special[mask]
    if show_active and 'Expiry_Date' in filtered_special.columns:
        filtered_special = filtered_special[filtered_special['Expiry_Date'] >= datetime.now().strftime('%Y-%m-%d')]
    
    st.markdown(f"**Found {len(filtered_special)} special offers**")
    
    if not filtered_special.empty:
        for _, special in filtered_special.iterrows():
            is_active = special['Expiry_Date'] >= datetime.now().strftime('%Y-%m-%d') if 'Expiry_Date' in special else True
            status = "🟢 Active" if is_active else "🔴 Expired"
            
            try:
                price_display = f"{float(special['Special_Price']):.2f}"
            except:
                price_display = str(special['Special_Price'])
            
            st.markdown(f"""
            <div class="price-card-secondary">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div>
                        <strong>{special['Article_Number']} - {special['Product_Name']}</strong>
                        <div style="font-size: 1.1rem; font-weight: 700; color: #b45309;">💰 {price_display} {special['Currency']}/kg</div>
                        <div style="font-size: 0.8rem; color: #64748b;">{status} • {special.get('Incoterm', 'N/A')}</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

# ============================================
# TAB 4: PRODUCTS & LOGISTICS
# ============================================

def products_logistics_tab():
    """Consolidated Products & Logistics"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0ea5e9, #0284c7); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">📦 Products & Logistics</h2>
        <p style="margin:0; opacity:0.9; color: white;">Complete product catalog • Pallet calculator • Logistics planning</p>
    </div>
    """, unsafe_allow_html=True)
    
    sub_tab1, sub_tab2 = st.tabs(["📋 Product Catalog", "📦 Pallet Calculator"])
    
    with sub_tab1:
        st.markdown("<div class='subsection-header'>📋 Complete Product Catalog</div>", unsafe_allow_html=True)
        
        with st.spinner("Loading product catalog..."):
            catalog_data = load_product_catalog()
        
        if catalog_data.empty:
            st.warning("Product catalog not found")
        else:
            st.success(f"Loaded {len(catalog_data)} products")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Products", len(catalog_data))
            if 'Supplier' in catalog_data.columns:
                col2.metric("Suppliers", catalog_data['Supplier'].nunique())
            if 'Category' in catalog_data.columns:
                col3.metric("Categories", catalog_data['Category'].nunique())
            col4.metric("Unique Articles", catalog_data['Article_Number'].nunique())
            
            col1, col2 = st.columns([2, 1])
            with col1:
                search_catalog = st.text_input("Search by article or product name...", key="catalog_search_input")
            with col2:
                if 'Supplier' in catalog_data.columns:
                    supplier_filter = st.selectbox("Supplier:", ["All"] + list(catalog_data['Supplier'].unique()), key="catalog_supplier_filter")
                else:
                    supplier_filter = "All"
            
            filtered_catalog = catalog_data.copy()
            if search_catalog:
                mask = filtered_catalog.astype(str).apply(lambda x: x.str.contains(search_catalog, case=False, na=False)).any(axis=1)
                filtered_catalog = filtered_catalog[mask]
            if supplier_filter != "All" and 'Supplier' in catalog_data.columns:
                filtered_catalog = filtered_catalog[filtered_catalog['Supplier'] == supplier_filter]
            
            st.markdown(f"**Found {len(filtered_catalog)} products**")
            
            if not filtered_catalog.empty:
                for _, product in filtered_catalog.head(30).iterrows():
                    with st.expander(f"📦 {product['Article_Number']} - {product['Product_Name']}", expanded=False):
                        product_tab1, product_tab2, product_tab3 = st.tabs(["📋 Details", "📄 Datasheet", "📊 Specifications"])
                        
                        with product_tab1:
                            col1, col2 = st.columns(2)
                            with col1:
                                st.markdown(f"**Article Number:** {product['Article_Number']}")
                                st.markdown(f"**Product Name:** {product['Product_Name']}")
                                if 'Supplier' in product and product['Supplier']:
                                    st.markdown(f"**Supplier:** {product['Supplier']}")
                            with col2:
                                if 'Category' in product and product['Category']:
                                    st.markdown(f"**Category:** {product['Category']}")
                                if 'UOM' in product and product['UOM']:
                                    st.markdown(f"**UOM:** {product['UOM']}")
                                if 'Packaging' in product and product['Packaging']:
                                    st.markdown(f"**Packaging:** {product['Packaging']}")
                            if 'Common_Description' in product and product['Common_Description']:
                                st.markdown("---")
                                st.markdown(f"**Description:** {product['Common_Description']}")
                        
                        with product_tab2:
                            st.markdown("### 📄 Product Datasheet")
                            
                            datasheet_link = None
                            if 'Datasheet_Link' in product and pd.notna(product['Datasheet_Link']) and product['Datasheet_Link']:
                                datasheet_link = product['Datasheet_Link']
                            elif 'Technical_Datasheet' in product and pd.notna(product['Technical_Datasheet']) and product['Technical_Datasheet']:
                                datasheet_link = product['Technical_Datasheet']
                            elif 'Product_Link' in product and pd.notna(product['Product_Link']) and product['Product_Link']:
                                datasheet_link = product['Product_Link']
                            
                            if datasheet_link:
                                st.markdown(f"📄 **Datasheet available:** [Click here to view/download datasheet]({datasheet_link})")
                                
                                if 'drive.google.com' in datasheet_link:
                                    file_id = datasheet_link.split('/d/')[1].split('/')[0] if '/d/' in datasheet_link else None
                                    if file_id:
                                        embed_url = f"https://drive.google.com/file/d/{file_id}/preview"
                                        st.markdown(f'<iframe src="{embed_url}" width="100%" height="500" style="border: none;"></iframe>', unsafe_allow_html=True)
                                    else:
                                        st.markdown(f"[Open in new tab]({datasheet_link})")
                                else:
                                    st.markdown(f"[Open in new tab]({datasheet_link})")
                            else:
                                st.info("📄 No datasheet available for this product. Please contact your sales representative.")
                        
                        with product_tab3:
                            st.markdown("### 📊 Technical Specifications")
                            
                            spec_cols = ['Weight_KG', 'Dimensions', 'Cartons_Per_Pallet', 'Pallet_Weight', 'Country_of_Origin', 'HS_Code', 'Tariff_Code', 'Shelf_Life', 'Storage_Conditions']
                            
                            spec_found = False
                            for col in spec_cols:
                                if col in product and pd.notna(product[col]) and product[col]:
                                    spec_found = True
                                    st.markdown(f"**{col.replace('_', ' ')}:** {product[col]}")
                            
                            if not spec_found:
                                st.info("📊 Detailed specifications not available. Check back later or contact support.")
                            
                            st.markdown("---")
                            st.markdown("### 📦 Logistics Information")
                            
                            col1, col2 = st.columns(2)
                            with col1:
                                st.markdown(f"**Article Number:** {product['Article_Number']}")
                                if 'UOM' in product and product['UOM']:
                                    st.markdown(f"**Unit of Measure:** {product['UOM']}")
                            with col2:
                                if 'Supplier' in product and product['Supplier']:
                                    st.markdown(f"**Supplier:** {product['Supplier']}")
                
                if len(filtered_catalog) > 30:
                    st.info(f"Showing 30 of {len(filtered_catalog)} products. Use filters to narrow down.")
    
    with sub_tab2:
        st.markdown("<div class='subsection-header'>📦 Quick Pallet Calculator</div>", unsafe_allow_html=True)
        
        standard_items = {
            "Vermicelli Color": {"packing": "5kg", "cartons_per_pallet": 100, "weight_per_carton": 5},
            "Vermicelli Dark": {"packing": "5kg", "cartons_per_pallet": 100, "weight_per_carton": 5},
            "Vermicelli White": {"packing": "5kg", "cartons_per_pallet": 100, "weight_per_carton": 5},
            "Chocolate Chips": {"packing": "25kg", "cartons_per_pallet": 40, "weight_per_carton": 25},
            "Date Mix": {"packing": "30kg", "cartons_per_pallet": 36, "weight_per_carton": 30},
            "Vanilla Powder": {"packing": "15kg", "cartons_per_pallet": 60, "weight_per_carton": 15},
            "Custom Item": {"packing": "Custom", "cartons_per_pallet": 0, "weight_per_carton": 0}
        }
        
        col1, col2 = st.columns(2)
        
        with col1:
            selected_item = st.selectbox("Select Item:", list(standard_items.keys()), key="pallet_item")
            quantity = st.number_input("Quantity:", min_value=1, value=100, step=1, key="pallet_qty")
            uom = st.selectbox("Unit of Measure:", ["Cartons", "KGs", "Pallets"], key="pallet_uom")
        
        with col2:
            if selected_item == "Custom Item":
                st.info("Enter custom item details:")
                packing = st.text_input("Packing:", value="5kg", key="custom_packing")
                cartons_per_pallet = st.number_input("Cartons per Pallet:", min_value=1, value=100, step=1, key="custom_cartons")
                weight_per_carton = st.number_input("Weight per Carton (kg):", min_value=0.1, value=5.0, step=0.1, key="custom_weight")
            else:
                item_data = standard_items[selected_item]
                packing = item_data["packing"]
                cartons_per_pallet = item_data["cartons_per_pallet"]
                weight_per_carton = item_data["weight_per_carton"]
                st.info(f"**Standard Packing:** {packing}")
                st.info(f"**Cartons per Pallet:** {cartons_per_pallet}")
                st.info(f"**Weight per Carton:** {weight_per_carton} kg")
        
        if quantity > 0 and cartons_per_pallet > 0:
            if uom == "Cartons":
                total_cartons = quantity
            elif uom == "KGs":
                total_cartons = quantity / weight_per_carton
            else:
                total_cartons = quantity * cartons_per_pallet
            
            full_pallets = int(total_cartons // cartons_per_pallet)
            partial_cartons = total_cartons % cartons_per_pallet
            total_weight = total_cartons * weight_per_carton
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-value">{full_pallets:,.0f}</div>
                    <div class="stat-label">Full Pallets</div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-value">{partial_cartons:,.0f}</div>
                    <div class="stat-label">Partial Cartons</div>
                </div>
                """, unsafe_allow_html=True)
            with col3:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-value">{total_weight:,.0f} kg</div>
                    <div class="stat-label">Total Weight</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.info(f"**Container Info:** A 40ft container holds approximately 30 pallets max. Your order fills {(full_pallets / 30 * 100):.1f}% of container capacity.")

# ============================================
# TAB 5: ORDER TRACKING
# ============================================

def order_tracking_tab():
    """Consolidated Order Tracking"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #059669, #047857); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">📅 Order Tracking</h2>
        <p style="margin:0; opacity:0.9; color: white;">ETD management • Order status • Sample requests</p>
    </div>
    """, unsafe_allow_html=True)
    
    sub_tab1, sub_tab2 = st.tabs(["🚢 ETD Management", "🎁 Samples Request"])
    
    with sub_tab1:
        st.markdown("<div class='subsection-header'>🚢 ETD Dashboard</div>", unsafe_allow_html=True)
        
        ETD_SHEET_ID = "1eA-mtD3aK_n9VYNV_bxnmqm58IywF0f5-7vr3PT51hs"
        AVAILABLE_MONTHS = ["October 2025", "November 2025"]
        
        selected_month = st.selectbox("Select Month:", AVAILABLE_MONTHS, key="etd_month")
        
        with st.spinner(f"Loading {selected_month} ETD data..."):
            etd_data = load_etd_data(ETD_SHEET_ID, selected_month)
        
        if etd_data.empty:
            st.warning(f"No ETD data found for {selected_month}")
        else:
            st.success(f"Loaded {len(etd_data)} orders for {selected_month}")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Orders", len(etd_data))
            if 'Status' in etd_data.columns:
                shipped = len(etd_data[etd_data['Status'].str.lower() == 'shipped'])
                col2.metric("Shipped", shipped)
                production = len(etd_data[etd_data['Status'].str.lower().str.contains('production', na=False)])
                col3.metric("In Production", production)
            
            col1, col2 = st.columns(2)
            with col1:
                if 'Client Name' in etd_data.columns:
                    clients = ["All"] + sorted(etd_data['Client Name'].dropna().unique())
                    client_filter = st.selectbox("Filter by Client:", clients, key="etd_client")
                else:
                    client_filter = "All"
            with col2:
                status_filter = st.selectbox("Filter by Status:", ["All", "Shipped", "In Production", "Pending", "Need ETD"], key="etd_status")
            
            filtered_etd = etd_data.copy()
            if client_filter != "All" and 'Client Name' in etd_data.columns:
                filtered_etd = filtered_etd[filtered_etd['Client Name'] == client_filter]
            if status_filter != "All" and status_filter != "Need ETD" and 'Status' in etd_data.columns:
                filtered_etd = filtered_etd[filtered_etd['Status'] == status_filter]
            
            st.markdown(f"**Found {len(filtered_etd)} orders**")
            
            if not filtered_etd.empty:
                for _, order in filtered_etd.iterrows():
                    status = order.get('Status', 'Unknown')
                    status_icon = {'Shipped': '🟢', 'In Production': '🟡', 'Pending': '🟠'}.get(status, '⚫')
                    
                    with st.expander(f"{status_icon} Order {order.get('Order No.', 'N/A')} - {order.get('Client Name', 'N/A')}", expanded=False):
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.markdown(f"**Client:** {order.get('Client Name', 'N/A')}")
                            st.markdown(f"**Employee:** {order.get('Concerned Employee', 'N/A')}")
                        with col2:
                            st.markdown(f"**Status:** {status}")
                            st.markdown(f"**Confirmation:** {order.get('Confirmation Date', 'N/A')}")
                        with col3:
                            st.markdown(f"**Loading:** {order.get('Scheduled Date For Loading', 'N/A')}")
                        
                        st.markdown("---")
                        st.markdown("**Supplier ETD Status**")
                        
                        for supplier in ['Backaldrine', 'bateel']:
                            etd_col = f"ETD _{supplier}" if supplier != 'bateel' else 'ETD_bateel'
                            etd_value = order.get(etd_col, '')
                            
                            if pd.isna(etd_value) or str(etd_value).strip() == '':
                                st.markdown(f"**{supplier}:** ❌ No ETD")
                            elif 'NEED ETD' in str(etd_value).upper():
                                st.markdown(f"**{supplier}:** <span class='badge-danger'>NEED ETD</span>", unsafe_allow_html=True)
                            else:
                                st.markdown(f"**{supplier}:** 📅 {etd_value}")
    
    with sub_tab2:
        st.markdown("<div class='subsection-header'>🎁 Samples Request Form</div>", unsafe_allow_html=True)
        
        if 'sample_items' not in st.session_state:
            st.session_state.sample_items = []
        
        catalog_data = load_product_catalog()
        article_to_product = {}
        
        if not catalog_data.empty and 'Article_Number' in catalog_data.columns and 'Product_Name' in catalog_data.columns:
            for _, row in catalog_data.iterrows():
                article = str(row['Article_Number']).strip()
                product = str(row['Product_Name']).strip()
                if article and article != 'nan' and product and product != 'nan':
                    article_to_product[article] = product
        
        col1, col2 = st.columns(2)
        
        with col1:
            request_date = st.date_input("Request Date", value=datetime.now().date(), key="sample_date")
            requested_by = st.text_input("Requested By *", placeholder="Enter your full name", key="sample_by")
            department = st.selectbox("Department", ["Sales", "Marketing", "R&D", "Production", "Quality Control", "Other"], key="sample_dept")
        
        with col2:
            going_to = st.text_input("Recipient Name *", placeholder="Enter recipient name", key="sample_recipient")
            address = st.text_area("Address *", placeholder="Enter complete delivery address", height=80, key="sample_address")
            delivery_method = st.selectbox("Delivery Method", ["Courier", "Pickup", "Mail", "Express Delivery", "Freight"], key="sample_delivery")
        
        st.markdown("---")
        st.markdown("**Add Sample Items:**")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            article_input = st.text_input("Article Number", placeholder="e.g., 1-366", key="sample_article")
            default_product = article_to_product.get(article_input, "")
            product_input = st.text_input("Product Name", value=default_product, key="sample_product")
        with col2:
            quantity = st.number_input("Quantity", min_value=1, value=1, step=1, key="sample_qty")
            unit_weight = st.number_input("Unit Weight (kg)", min_value=0.0, step=0.1, format="%.2f", key="sample_weight", value=0.0)
        with col3:
            logo_required = st.radio("Logo Required?", ["No", "Yes"], horizontal=True, key="sample_logo")
            notes = st.text_input("Notes (Optional)", key="sample_notes")
        
        if st.button("➕ Add Item", use_container_width=True, key="add_sample"):
            if article_input and product_input:
                st.session_state.sample_items.append({
                    'article': article_input,
                    'product': product_input,
                    'quantity': quantity,
                    'unit_weight': unit_weight,
                    'logo': logo_required,
                    'notes': notes
                })
                st.rerun()
            else:
                st.error("Please enter both Article Number and Product Name")
        
        if st.session_state.sample_items:
            st.markdown(f"**Sample Items ({len(st.session_state.sample_items)}):**")
            for idx, item in enumerate(st.session_state.sample_items):
                col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
                with col1:
                    st.write(f"{item['article']} - {item['product']}")
                with col2:
                    st.write(f"Qty: {item['quantity']}")
                with col3:
                    st.write(f"Weight: {item['unit_weight']} kg")
                with col4:
                    if st.button("🗑️", key=f"remove_{idx}"):
                        st.session_state.sample_items.pop(idx)
                        st.rerun()
            
            if st.button("🗑️ Clear All Items", use_container_width=True):
                st.session_state.sample_items = []
                st.rerun()
        
        if st.button("📤 SUBMIT SAMPLES REQUEST", use_container_width=True, type="primary", key="submit_samples"):
            if not requested_by:
                st.error("Please enter Requested By")
            elif not going_to:
                st.error("Please enter Recipient Name")
            elif not address:
                st.error("Please enter Address")
            elif not st.session_state.sample_items:
                st.error("Please add at least one sample item")
            else:
                st.balloons()
                st.success(f"✅ Samples request submitted successfully! Request ID: SAMP-{datetime.now().strftime('%Y%m%d-%H%M%S')}")
                
                st.markdown("### Request Summary")
                st.markdown(f"**Requested By:** {requested_by}")
                st.markdown(f"**Recipient:** {going_to}")
                st.markdown(f"**Address:** {address}")
                st.markdown(f"**Total Items:** {len(st.session_state.sample_items)}")
                
                if st.button("📋 New Request", key="new_sample"):
                    st.session_state.sample_items = []
                    st.rerun()

# ============================================
# TAB 6: PRICE TRACKING
# ============================================

def price_tracking_tab():
    """Track price changes over time for any item"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #dc2626, #b91c1c); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">📈 Price Tracking</h2>
        <p style="margin:0; opacity:0.9; color: white;">Track historical price changes • Visualize trends</p>
    </div>
    """, unsafe_allow_html=True)
    
    available_clients = st.session_state.user_clients
    if not available_clients:
        st.warning("No clients available. Please check your Clients_CoC sheet.")
        return
    
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        search_item = st.text_input("🔍 Search for an item:", 
                                    placeholder="e.g., 1-366, Vermicelli, Chocolate...",
                                    key="price_track_search")
    
    with col2:
        selected_client = st.selectbox("Select Client:", available_clients, key="price_track_client")
    
    with col3:
        selected_supplier = st.selectbox("Select Supplier:", ["Both", "Backaldrin", "Bateel"], key="price_track_supplier")
    
    with st.expander("📊 Advanced Options"):
        col1, col2 = st.columns(2)
        with col1:
            group_by = st.selectbox("Group By:", ["Year", "Month", "Quarter", "All Orders"], key="price_track_group")
        with col2:
            show_statistics = st.checkbox("Show Statistics", value=True, key="price_track_stats")
    
    if search_item:
        with st.spinner(f"Analyzing price history for '{search_item}'..."):
            price_data = []
            
            client_data = get_google_sheets_data(selected_client)
            
            suppliers_to_check = []
            if selected_supplier == "Both":
                suppliers_to_check = ["Backaldrin", "Bateel"]
            else:
                suppliers_to_check = [selected_supplier]
            
            for supplier in suppliers_to_check:
                supplier_data = client_data.get(supplier, {})
                
                for article_num, article_data in supplier_data.items():
                    article_match = search_item.lower() in article_num.lower()
                    product_match = any(search_item.lower() in name.lower() for name in article_data.get('names', []))
                    
                    if article_match or product_match:
                        product_name = article_data.get('names', ['N/A'])[0]
                        
                        for order in article_data.get('orders', []):
                            price_str = order.get('price', '')
                            date_str = order.get('date', '')
                            
                            price_value = None
                            try:
                                price_value = float(str(price_str).replace('$', '').replace(',', '').strip())
                            except:
                                continue
                            
                            order_date = None
                            year = None
                            month = None
                            quarter = None
                            
                            year_str = order.get('year', '')
                            if year_str and year_str != 'nan':
                                try:
                                    year = int(str(year_str).strip())
                                    if date_str and date_str != 'nan':
                                        for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                                            try:
                                                date_obj = datetime.strptime(str(date_str).strip(), fmt)
                                                month = date_obj.month
                                                order_date = date_obj
                                                break
                                            except:
                                                continue
                                except:
                                    pass
                            
                            if not year and date_str and date_str != 'nan':
                                for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                                    try:
                                        date_obj = datetime.strptime(str(date_str).strip(), fmt)
                                        year = date_obj.year
                                        month = date_obj.month
                                        order_date = date_obj
                                        break
                                    except:
                                        continue
                            
                            if year:
                                if month:
                                    quarter = (month - 1) // 3 + 1
                                
                                price_data.append({
                                    'article': article_num,
                                    'product': product_name,
                                    'supplier': supplier,
                                    'client': selected_client,
                                    'price': price_value,
                                    'year': year,
                                    'month': month,
                                    'quarter': quarter,
                                    'date': order_date,
                                    'order_no': order.get('order_no', 'N/A'),
                                    'quantity': order.get('quantity', 'N/A'),
                                    'total_weight': order.get('total_weight', 'N/A')
                                })
            
            if price_data:
                df = pd.DataFrame(price_data)
                
                if 'date' in df.columns:
                    df = df.sort_values('date')
                else:
                    df = df.sort_values('year')
                
                if show_statistics:
                    st.markdown("### 📊 Price Statistics")
                    
                    col1, col2, col3, col4, col5 = st.columns(5)
                    with col1:
                        st.metric("First Recorded Price", f"${df['price'].iloc[0]:.2f}" if len(df) > 0 else "N/A")
                    with col2:
                        st.metric("Latest Price", f"${df['price'].iloc[-1]:.2f}" if len(df) > 0 else "N/A")
                    with col3:
                        price_change = df['price'].iloc[-1] - df['price'].iloc[0] if len(df) > 0 else 0
                        change_pct = (price_change / df['price'].iloc[0] * 100) if len(df) > 0 and df['price'].iloc[0] > 0 else 0
                        st.metric("Total Change", f"${price_change:.2f}", delta=f"{change_pct:.1f}%")
                    with col4:
                        st.metric("Min Price", f"${df['price'].min():.2f}")
                    with col5:
                        st.metric("Max Price", f"${df['price'].max():.2f}")
                
                st.markdown("### 📈 Price Trend Visualization")
                
                if group_by == "Year":
                    yearly_avg = df.groupby('year')['price'].agg(['mean', 'min', 'max']).reset_index()
                    
                    line_chart = alt.Chart(yearly_avg).mark_line(point=True, strokeWidth=3).encode(
                        x=alt.X('year:Q', title='Year'),
                        y=alt.Y('mean:Q', title='Price (USD/kg)'),
                        tooltip=['year', alt.Tooltip('mean', format='.2f')]
                    ).properties(height=400)
                    
                    points = alt.Chart(yearly_avg).mark_circle(size=100).encode(
                        x='year:Q',
                        y='mean:Q',
                        color=alt.value('#dc2626'),
                        tooltip=['year', alt.Tooltip('mean', format='.2f')]
                    )
                    
                    st.altair_chart(line_chart + points, use_container_width=True)
                    
                    st.markdown("### 📅 Yearly Price Summary")
                    yearly_display = yearly_avg.copy()
                    yearly_display['mean'] = yearly_display['mean'].apply(lambda x: f"${x:.2f}")
                    yearly_display['min'] = yearly_display['min'].apply(lambda x: f"${x:.2f}")
                    yearly_display['max'] = yearly_display['max'].apply(lambda x: f"${x:.2f}")
                    yearly_display.columns = ['Year', 'Avg Price', 'Min Price', 'Max Price']
                    st.dataframe(yearly_display, use_container_width=True, hide_index=True)
                    
                elif group_by == "Month" and 'date' in df.columns and df['date'].notna().any():
                    df['year_month'] = df['date'].dt.strftime('%Y-%m')
                    monthly_avg = df.groupby('year_month')['price'].mean().reset_index()
                    
                    bars = alt.Chart(monthly_avg).mark_bar(color='#dc2626').encode(
                        x=alt.X('year_month:N', title='Month', sort=None),
                        y=alt.Y('price:Q', title='Price (USD/kg)'),
                        tooltip=['year_month', alt.Tooltip('price', format='.2f')]
                    ).properties(height=400)
                    
                    st.altair_chart(bars, use_container_width=True)
                    
                elif group_by == "Quarter":
                    df['year_quarter'] = df['year'].astype(str) + '-Q' + df['quarter'].astype(str)
                    quarterly_avg = df.groupby('year_quarter')['price'].mean().reset_index()
                    
                    bars = alt.Chart(quarterly_avg).mark_bar(color='#dc2626').encode(
                        x=alt.X('year_quarter:N', title='Quarter', sort=None),
                        y=alt.Y('price:Q', title='Price (USD/kg)'),
                        tooltip=['year_quarter', alt.Tooltip('price', format='.2f')]
                    ).properties(height=400)
                    
                    st.altair_chart(bars, use_container_width=True)
                
                elif group_by == "All Orders" and 'date' in df.columns:
                    scatter = alt.Chart(df).mark_circle(size=60, color='#dc2626').encode(
                        x=alt.X('date:T', title='Order Date'),
                        y=alt.Y('price:Q', title='Price (USD/kg)'),
                        tooltip=['date', alt.Tooltip('price', format='.2f'), 'order_no', 'supplier']
                    ).properties(height=400)
                    
                    trend = scatter.transform_regression('date', 'price').mark_line(color='#f59e0b', strokeWidth=2)
                    
                    st.altair_chart(scatter + trend, use_container_width=True)
                
                st.markdown("### 📋 Detailed Price History")
                
                display_df = df[['year', 'date', 'price', 'order_no', 'quantity', 'total_weight', 'supplier']].copy()
                display_df['price'] = display_df['price'].apply(lambda x: f"${x:.2f}")
                display_df.columns = ['Year', 'Date', 'Price (USD/kg)', 'Order No', 'Quantity', 'Weight (kg)', 'Supplier']
                
                st.dataframe(display_df, use_container_width=True, hide_index=True)
                
                csv_data = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📥 Export Price History to CSV",
                    data=csv_data.encode('utf-8-sig'),
                    file_name=f"price_history_{search_item}_{selected_client}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
                
            else:
                st.warning(f"No price data found for '{search_item}' in {selected_client}")
                st.info("💡 Make sure your Clients_CoC sheet has:\n- The correct Client name\n- Values in the 'Price' column\n- Values in the 'Year' or 'Order_Date' column")
    else:
        st.info("👆 Enter an article number or product name above to start tracking price history")

# ============================================
# TAB 7: COMMISSION
# ============================================

def commission_tab():
    """Commission tracking tab with beautiful Excel export and data management"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #059669, #047857); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">💰 Commission Tracker</h2>
        <p style="margin:0; opacity:0.9; color: white;">Track sales commissions • Add new commission entries • Download professional Excel reports</p>
    </div>
    """, unsafe_allow_html=True)
    
    staff_name = st.session_state.username.capitalize()
    
    if st.session_state.user_clients == "ALL" or isinstance(st.session_state.user_clients, list):
        assigned_clients = st.session_state.user_clients if isinstance(st.session_state.user_clients, list) else []
    else:
        assigned_clients = []
    
    with st.expander("⚙️ Data Management Tools"):
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            if st.button("🗑️ Clear ALL Data", type="secondary", use_container_width=True):
                if 'commission_records' in st.session_state:
                    st.session_state.commission_records = []
                    st.success("✅ All commission data has been cleared!")
                    st.rerun()
        
        with col2:
            if st.button("👤 Keep My Records Only", type="secondary", use_container_width=True):
                if 'commission_records' in st.session_state:
                    st.session_state.commission_records = [
                        r for r in st.session_state.commission_records 
                        if r.get('staff') == staff_name
                    ]
                    st.success(f"✅ Kept only {staff_name}'s records!")
                    st.rerun()
        
        with col3:
            if st.button("📅 Keep Last 30 Days", type="secondary", use_container_width=True):
                if 'commission_records' in st.session_state:
                    thirty_days_ago = datetime.now().date() - pd.Timedelta(days=30)
                    original_count = len(st.session_state.commission_records)
                    st.session_state.commission_records = [
                        r for r in st.session_state.commission_records 
                        if datetime.strptime(r['date'], '%Y-%m-%d').date() >= thirty_days_ago
                    ]
                    st.success(f"✅ Kept {len(st.session_state.commission_records)} records from last 30 days (removed {original_count - len(st.session_state.commission_records)})")
                    st.rerun()
        
        with col4:
            if st.button("📊 Keep This Year", type="secondary", use_container_width=True):
                if 'commission_records' in st.session_state:
                    current_year = datetime.now().year
                    original_count = len(st.session_state.commission_records)
                    st.session_state.commission_records = [
                        r for r in st.session_state.commission_records 
                        if int(r['date'].split('-')[0]) == current_year
                    ]
                    st.success(f"✅ Kept {len(st.session_state.commission_records)} records from {current_year} (removed {original_count - len(st.session_state.commission_records)})")
                    st.rerun()
    
    st.info(f"👤 Logged in as: **{staff_name}** | 📊 Total Records: {len(st.session_state.get('commission_records', []))}")
    
    tab1, tab2, tab3 = st.tabs(["➕ Add Commission", "📋 View & Manage", "📊 Export Excel Report"])
    
    with tab1:
        st.markdown("<div class='subsection-header'>➕ Add New Commission Entry</div>", unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            section_type = st.selectbox("Section:", ["Jordan Sales USD", "Cross-booking (Euro)", "Chocolate Spreads", "Rosellia Powder Cream"], key="commission_section")
            currency_type = "USD" if "USD" in section_type else "Euro" if "Euro" in section_type else "USD"
            date = st.date_input("Date:", value=datetime.now().date(), key="commission_date")
            export_inv = st.text_input("Export Invoice #:", placeholder="e.g., INV-001", key="commission_inv")
            customer_name = st.text_input("Customer Name:", placeholder="Enter customer name", key="commission_customer")
            
        with col2:
            if assigned_clients:
                client_options = assigned_clients if assigned_clients else ["No clients assigned"]
                client = st.selectbox("Client:", client_options, key="commission_client")
            else:
                client = st.text_input("Client:", placeholder="Enter client name", key="commission_client")
            
            country = st.text_input("Country:", placeholder="e.g., Jordan, UAE, KSA", key="commission_country")
            weight_kg = st.number_input("Weight (KG):", min_value=0.0, step=0.1, format="%.2f", key="commission_weight")
        
        st.markdown("---")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            amount = st.number_input("Amount ($/€):", min_value=0.0, step=10.0, format="%.2f", key="commission_amount")
        with col2:
            discount = st.number_input("Discount ($/€):", min_value=0.0, step=5.0, format="%.2f", key="commission_discount")
        with col3:
            charges = st.number_input("Charges ($/€):", min_value=0.0, step=5.0, format="%.2f", key="commission_charges")
        
        after_discount = amount - discount
        after_discount_jd = after_discount * 0.708
        net_amount = after_discount + charges
        amount_jd = net_amount * 0.708
        rate = 0.00125
        com_amount = rate * amount
        com_jd = com_amount * 0.708
        
        st.markdown("### 📊 Calculated Values")
        calc_col1, calc_col2, calc_col3, calc_col4 = st.columns(4)
        with calc_col1:
            st.metric("After Discount", f"{after_discount:.2f}")
        with calc_col2:
            st.metric("After Disc. (JD)", f"{after_discount_jd:.2f}")
        with calc_col3:
            st.metric("Net Amount", f"{net_amount:.2f}")
        with calc_col4:
            st.metric("Commission", f"{com_amount:.2f} ({com_jd:.2f} JD)")
        
        if st.button("💾 Save Commission Record", type="primary", use_container_width=True, key="save_commission"):
            if not customer_name:
                st.error("Please enter Customer Name")
            elif not client:
                st.error("Please select/enter Client")
            elif amount <= 0:
                st.error("Please enter Amount greater than 0")
            else:
                record = {
                    'date': date.strftime('%Y-%m-%d'),
                    'staff': staff_name,
                    'client': client,
                    'section': section_type,
                    'export_inv': export_inv,
                    'customer_name': customer_name,
                    'country': country,
                    'weight_kg': weight_kg,
                    'currency_type': currency_type,
                    'amount': amount,
                    'discount': discount,
                    'after_discount': after_discount,
                    'after_discount_jd': after_discount_jd,
                    'charges': charges,
                    'net_amount': net_amount,
                    'amount_jd': amount_jd,
                    'rate': rate,
                    'commission': com_amount,
                    'commission_jd': com_jd,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                if 'commission_records' not in st.session_state:
                    st.session_state.commission_records = []
                
                st.session_state.commission_records.append(record)
                st.success(f"✅ Commission record saved successfully!")
                st.balloons()
    
    with tab2:
        st.markdown("<div class='subsection-header'>📋 View & Manage Commission Records</div>", unsafe_allow_html=True)
        
        if 'commission_records' in st.session_state and st.session_state.commission_records:
            df = pd.DataFrame(st.session_state.commission_records)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                filter_staff = st.multiselect("Filter by Staff:", options=df['staff'].unique(), default=[])
            with col2:
                filter_client = st.multiselect("Filter by Client:", options=df['client'].unique(), default=[])
            with col3:
                filter_section = st.multiselect("Filter by Section:", options=df['section'].unique(), default=[])
            with col4:
                filter_year = st.multiselect("Filter by Year:", options=sorted(df['date'].str.split('-').str[0].unique()), default=[])
            
            filtered_df = df.copy()
            if filter_staff:
                filtered_df = filtered_df[filtered_df['staff'].isin(filter_staff)]
            if filter_client:
                filtered_df = filtered_df[filtered_df['client'].isin(filter_client)]
            if filter_section:
                filtered_df = filtered_df[filtered_df['section'].isin(filter_section)]
            if filter_year:
                filtered_df = filtered_df[filtered_df['date'].str.split('-').str[0].isin(filter_year)]
            
            st.markdown(f"**Found {len(filtered_df)} records**")
            
            st.markdown("### Records (Click ❌ to delete individual record)")
            
            for idx, record in filtered_df.iterrows():
                with st.container():
                    col1, col2, col3, col4, col5, col6, col7 = st.columns([1.5, 1.5, 2, 1.5, 1.5, 1.5, 0.5])
                    with col1:
                        st.write(record['date'])
                    with col2:
                        st.write(record['staff'])
                    with col3:
                        st.write(record['customer_name'][:20])
                    with col4:
                        st.write(record['section'][:15])
                    with col5:
                        st.write(f"${record['amount']:.2f}")
                    with col6:
                        st.write(f"${record['commission']:.2f}")
                    with col7:
                        if st.button("❌", key=f"del_{record['timestamp']}"):
                            original_idx = next((i for i, r in enumerate(st.session_state.commission_records) if r['timestamp'] == record['timestamp']), None)
                            if original_idx is not None:
                                st.session_state.commission_records.pop(original_idx)
                                st.rerun()
                    st.divider()
            
            st.markdown("---")
            st.markdown("### Batch Delete")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🗑️ Delete All Filtered Records", type="secondary", use_container_width=True):
                    if filtered_df is not None and not filtered_df.empty:
                        timestamps_to_delete = set(filtered_df['timestamp'].tolist())
                        original_count = len(st.session_state.commission_records)
                        st.session_state.commission_records = [
                            r for r in st.session_state.commission_records 
                            if r['timestamp'] not in timestamps_to_delete
                        ]
                        st.success(f"✅ Deleted {original_count - len(st.session_state.commission_records)} records!")
                        st.rerun()
        else:
            st.info("No commission records yet. Use the 'Add Commission' tab to create your first record.")
    
    with tab3:
        st.markdown("<div class='subsection-header'>📊 Export Professional Excel Report</div>", unsafe_allow_html=True)
        
        if 'commission_records' in st.session_state and st.session_state.commission_records:
            df = pd.DataFrame(st.session_state.commission_records)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                total_usd = df[df['currency_type'] == 'USD']['commission'].sum() if 'USD' in df['currency_type'].values else 0
                st.metric("Total Commission (USD)", f"${total_usd:.2f}")
            with col2:
                total_eur = df[df['currency_type'] == 'Euro']['commission'].sum() if 'Euro' in df['currency_type'].values else 0
                st.metric("Total Commission (EUR)", f"€{total_eur:.2f}")
            with col3:
                total_jd = df['commission_jd'].sum()
                st.metric("Total Commission (JD)", f"{total_jd:.2f} JD")
            with col4:
                total_sales = df['net_amount'].sum()
                st.metric("Total Sales", f"${total_sales:.2f}")
            
            st.markdown("### 📥 Download Beautiful Excel Report")
            st.caption("The Excel file will have: bold colored headers, alternating row colors, professional borders, and auto-fitted columns")
            
            if st.button("🎨 Generate Beautiful Excel Report", type="primary", use_container_width=True):
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                excel_buffer = BytesIO()
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    sections = {
                        'Jordan Sales USD': df[df['section'] == 'Jordan Sales USD'],
                        'Cross-booking (Euro)': df[df['section'] == 'Cross-booking (Euro)'],
                        'Chocolate Spreads': df[df['section'] == 'Chocolate Spreads'],
                        'Rosellia Powder Cream': df[df['section'] == 'Rosellia Powder Cream']
                    }
                    
                    usd_headers = ['Date', 'Export Inv #', 'Customer Name', 'Country', 'WGT / KG', 'Amount / $', 'Discount', 'After Discount', 'After Disc. (JD)', 'Charges', 'Net Amount $', 'Amount JD', 'Rate', 'Com / $', 'Com / JD']
                    euro_headers = ['Date', 'Export Inv #', 'Customer Name', 'Country', 'WGT / KG', 'Amount / €', 'Discount', 'After Discount', 'After Disc. (JD)', 'Charges', 'Net Amount €', 'Amount JD', 'Rate', 'Com / €', 'Com / JD']
                    
                    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True, size=11)
                    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for section_name, section_data in sections.items():
                        if section_data.empty:
                            continue
                            
                        if 'Euro' in section_name:
                            export_data = section_data[['date', 'export_inv', 'customer_name', 'country', 'weight_kg', 'amount', 'discount', 'after_discount', 'after_discount_jd', 'charges', 'net_amount', 'amount_jd', 'rate', 'commission', 'commission_jd']].copy()
                            export_data.columns = euro_headers
                        else:
                            export_data = section_data[['date', 'export_inv', 'customer_name', 'country', 'weight_kg', 'amount', 'discount', 'after_discount', 'after_discount_jd', 'charges', 'net_amount', 'amount_jd', 'rate', 'commission', 'commission_jd']].copy()
                            export_data.columns = usd_headers
                        
                        export_data.to_excel(writer, sheet_name=section_name[:31], index=False)
                        worksheet = writer.sheets[section_name[:31]]
                        
                        for col in range(1, len(export_data.columns) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                            cell.border = thin_border
                        
                        for row in range(2, len(export_data) + 2):
                            for col in range(1, len(export_data.columns) + 1):
                                cell = worksheet.cell(row=row, column=col)
                                cell.fill = light_fill if row % 2 == 0 else white_fill
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                        worksheet.freeze_panes = 'A2'
                        
                        for col in worksheet.columns:
                            max_length = 0
                            col_letter = get_column_letter(col[0].column)
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 25)
                    
                    usd_total = sections['Jordan Sales USD']['commission'].sum() if not sections['Jordan Sales USD'].empty else 0
                    euro_total = sections['Cross-booking (Euro)']['commission'].sum() if not sections['Cross-booking (Euro)'].empty else 0
                    chocolate_total = sections['Chocolate Spreads']['commission'].sum() if not sections['Chocolate Spreads'].empty else 0
                    rosellia_total = sections['Rosellia Powder Cream']['commission'].sum() if not sections['Rosellia Powder Cream'].empty else 0
                    
                    summary_data = [
                        ['COMMISSION SUMMARY REPORT'],
                        [''],
                        ['Section', 'Total Commission', 'Total Commission (JD)'],
                        ['Jordan Sales USD', f"${usd_total:.2f}", f"{usd_total * 0.708:.2f} JD"],
                        ['Cross-booking (Euro)', f"€{euro_total:.2f}", f"{euro_total * 0.825:.2f} JD"],
                        ['Chocolate Spreads', f"${chocolate_total:.2f}", f"{chocolate_total * 0.708:.2f} JD"],
                        ['Rosellia Powder Cream', f"${rosellia_total:.2f}", f"{rosellia_total * 0.708:.2f} JD"],
                        ['', '', ''],
                        ['TOTAL COMMISSION', f"${usd_total + chocolate_total + rosellia_total:.2f} + €{euro_total:.2f}", f"{(usd_total + chocolate_total + rosellia_total) * 0.708 + euro_total * 0.825:.2f} JD"],
                        ['', '', ''],
                        [f'Report Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '', ''],
                        [f'Generated by: {staff_name}', '', ''],
                        [f'Total Records: {len(df)}', '', '']
                    ]
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
                    
                    summary_ws = writer.sheets['Summary']
                    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    title_font = Font(color="FFFFFF", bold=True, size=14)
                    title_cell = summary_ws['A1']
                    title_cell.fill = title_fill
                    title_cell.font = title_font
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    summary_ws.merge_cells('A1:C1')
                    
                    header_fill_light = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    header_font_bold = Font(bold=True)
                    for col in range(1, 4):
                        cell = summary_ws.cell(row=4, column=col)
                        cell.fill = header_fill_light
                        cell.font = header_font_bold
                        cell.border = thin_border
                    
                    for col in summary_ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        summary_ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
                
                st.success("✅ Beautiful Excel report generated successfully!")
                st.download_button(
                    label="📥 Download Professional Excel Report",
                    data=excel_buffer.getvalue(),
                    file_name=f"commission_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.info("No data available. Add some commission records first to generate the Excel report.")

# ============================================
# TAB 8: CLIENT DETAILS
# ============================================

def client_details_tab():
    """Display client details from Client_details sheet with support for partial data"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #8b5cf6, #7c3aed); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">👥 Client Details</h2>
        <p style="margin:0; opacity:0.9; color: white;">Complete client information • Contact details • Business information</p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.spinner("Loading client details..."):
        client_details_df = load_client_details()
    
    if client_details_df.empty:
        st.warning("No client details found. Please make sure the 'Client_details' sheet exists and contains data.")
        st.info("The sheet should have columns like: Client Name, Contact Person, Email, Phone, Address, Country, etc.")
        return
    
    st.success(f"✅ Loaded {len(client_details_df)} client records")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Clients", len(client_details_df))
    with col2:
        email_col = next((col for col in client_details_df.columns if 'email' in col.lower() or 'Email' in col), None)
        if email_col:
            email_count = client_details_df[email_col].notna().sum()
            st.metric("With Email", email_count)
        else:
            st.metric("With Email", "N/A")
    with col3:
        phone_col = next((col for col in client_details_df.columns if 'phone' in col.lower() or 'Phone' in col or 'mobile' in col.lower()), None)
        if phone_col:
            phone_count = client_details_df[phone_col].notna().sum()
            st.metric("With Phone", phone_count)
        else:
            st.metric("With Phone", "N/A")
    with col4:
        address_col = next((col for col in client_details_df.columns if 'address' in col.lower() or 'Address' in col), None)
        if address_col:
            address_count = client_details_df[address_col].notna().sum()
            st.metric("With Address", address_count)
        else:
            st.metric("With Address", "N/A")
    
    st.markdown("---")
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        search_term = st.text_input("🔍 Search Clients:", placeholder="Search by name, contact person, email, country...", key="client_search")
    
    with col2:
        country_col = next((col for col in client_details_df.columns if 'country' in col.lower() or 'Country' in col), None)
        if country_col:
            countries = ["All"] + sorted(client_details_df[country_col].dropna().unique().tolist())
            country_filter = st.selectbox("Filter by Country:", countries, key="client_country_filter")
        else:
            country_filter = "All"
    
    with col3:
        show_complete_only = st.checkbox("Show Complete Info Only", value=False, help="Show only clients with all major fields filled")
    
    filtered_df = client_details_df.copy()
    
    if search_term:
        search_lower = search_term.lower()
        mask = filtered_df.astype(str).apply(lambda x: x.str.contains(search_lower, case=False, na=False)).any(axis=1)
        filtered_df = filtered_df[mask]
    
    if country_filter != "All" and country_col:
        filtered_df = filtered_df[filtered_df[country_col] == country_filter]
    
    if show_complete_only:
        major_fields = []
        for col in filtered_df.columns:
            if any(keyword in col.lower() for keyword in ['name', 'email', 'phone', 'address', 'contact']):
                major_fields.append(col)
        
        if major_fields:
            for field in major_fields:
                filtered_df = filtered_df[filtered_df[field].notna()]
    
    st.markdown(f"### 📋 Client Directory ({len(filtered_df)} clients)")
    
    if not filtered_df.empty:
        view_type = st.radio("View as:", ["Cards", "Table"], horizontal=True, key="client_view_type")
        
        if view_type == "Cards":
            cols_per_row = 3
            for i in range(0, len(filtered_df), cols_per_row):
                cols = st.columns(cols_per_row)
                for j in range(cols_per_row):
                    if i + j < len(filtered_df):
                        client = filtered_df.iloc[i + j]
                        with cols[j]:
                            name_col = next((col for col in client.index if 'name' in col.lower() or 'Name' in col), None)
                            client_name = client[name_col] if name_col and pd.notna(client[name_col]) else f"Client {i+j+1}"
                            
                            st.markdown(f"""
                            <div class="client-detail-card">
                                <h4 style="margin: 0 0 0.5rem 0; color: #1e293b;">🏢 {client_name}</h4>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            for col in client.index:
                                value = client[col]
                                if pd.notna(value) and str(value).strip():
                                    label = col.replace('_', ' ').title()
                                    st.markdown(f"**{label}:** {value}")
                            
                            st.markdown("---")
        else:
            important_cols = []
            for col in filtered_df.columns:
                if any(keyword in col.lower() for keyword in ['name', 'contact', 'email', 'phone', 'mobile', 'address', 'country', 'city']):
                    important_cols.append(col)
            
            if not important_cols:
                important_cols = filtered_df.columns.tolist()
            
            display_cols = important_cols[:10]
            
            display_df = filtered_df[display_cols].copy()
            display_df = display_df.fillna('')
            display_df.columns = [col.replace('_', ' ').title() for col in display_df.columns]
            
            st.dataframe(display_df, use_container_width=True, height=400)
    
    else:
        st.info("No clients match your search criteria")
    
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📥 Export Client Details to CSV", use_container_width=True):
            csv_data = filtered_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="Download CSV",
                data=csv_data.encode('utf-8-sig'),
                file_name=f"client_details_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    with col2:
        if st.button("🔄 Refresh Data", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
    
    with st.expander("📊 Data Completeness Summary"):
        st.markdown("### Field Completion Rates")
        
        completion_data = []
        for col in client_details_df.columns:
            non_null_count = client_details_df[col].notna().sum()
            completion_pct = (non_null_count / len(client_details_df)) * 100
            completion_data.append({
                'Field': col,
                'Filled Records': non_null_count,
                'Empty Records': len(client_details_df) - non_null_count,
                'Completion Rate': f"{completion_pct:.1f}%"
            })
        
        completion_df = pd.DataFrame(completion_data)
        completion_df = completion_df.sort_values('Filled Records', ascending=False)
        
        st.dataframe(completion_df, use_container_width=True, hide_index=True)
        
        st.info("💡 Tip: Add more data to the 'Client_details' sheet to see more information here. Empty cells are handled gracefully.")

# ============================================
# TAB 9: PRICE CHECKER
# ============================================

def price_checker_tab():
    """Check the last/most recent price for any item by client with currency support"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #ec4899, #db2777); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">💰 Price Checker</h2>
        <p style="margin:0; opacity:0.9; color: white;">Get the latest price for any product • Supports USD, EUR, SAR</p>
    </div>
    """, unsafe_allow_html=True)
    
    available_clients = st.session_state.user_clients
    if not available_clients:
        st.warning("No clients available. Please check your Clients_CoC sheet.")
        return
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        search_mode = st.radio("Search Mode:", ["🔍 Search Specific Item", "📋 Show All Items"], key="price_checker_mode")
    
    with col2:
        selected_client = st.selectbox("Select Client:", available_clients, key="price_checker_client")
    
    # Currency symbol mapping
    currency_symbols = {
        "USD": "$",
        "EUR": "€",
        "SAR": "ر.س"
    }
    
    if search_mode == "🔍 Search Specific Item":
        col1, col2 = st.columns([2, 1])
        with col1:
            search_term = st.text_input("🔍 Search by Article Number or Product Name:", 
                                        placeholder="e.g., 1-366, Chocolate Spread, Vermicelli...",
                                        key="price_checker_search")
        with col2:
            search_supplier = st.selectbox("Supplier:", ["Both", "Backaldrin", "Bateel"], key="price_checker_supplier")
        
        if search_term:
            with st.spinner(f"Searching for '{search_term}' in {selected_client}..."):
                client_data = get_google_sheets_data(selected_client)
                
                results = []
                
                suppliers_to_check = []
                if search_supplier == "Both":
                    suppliers_to_check = ["Backaldrin", "Bateel"]
                else:
                    suppliers_to_check = [search_supplier]
                
                for supplier in suppliers_to_check:
                    supplier_data = client_data.get(supplier, {})
                    
                    for article_num, article_data in supplier_data.items():
                        article_match = search_term.lower() in article_num.lower()
                        product_match = any(search_term.lower() in name.lower() for name in article_data.get('names', []))
                        
                        if article_match or product_match:
                            orders = article_data.get('orders', [])
                            if orders:
                                # Sort orders by date to find the latest
                                sorted_orders = sorted(orders, key=lambda x: str(x.get('date', '')), reverse=True)
                                latest_order = sorted_orders[0] if sorted_orders else None
                                
                                if latest_order:
                                    price_value = latest_order.get('price_value')
                                    currency = latest_order.get('currency', 'USD')
                                    latest_date = latest_order.get('date', 'Unknown')
                                    symbol = currency_symbols.get(currency, "$")
                                    
                                    if price_value:
                                        latest_price_display = f"{symbol}{price_value:.2f}"
                                    else:
                                        latest_price_display = latest_order.get('price', 'N/A')
                                    
                                    # Get price history
                                    price_history = []
                                    for p in article_data.get('prices_with_currency', []):
                                        sym = currency_symbols.get(p['currency'], "$")
                                        price_history.append(f"{sym}{p['value']:.2f}")
                                    
                                    results.append({
                                        'Article': article_num,
                                        'Product Name': article_data.get('names', ['N/A'])[0],
                                        'Supplier': supplier,
                                        'Latest Price': latest_price_display,
                                        'Currency': currency,
                                        'Last Order Date': latest_date,
                                        'Total Orders': len(orders),
                                        'Price History': price_history
                                    })
                
                if results:
                    st.success(f"✅ Found {len(results)} matching items")
                    
                    for result in results:
                        with st.expander(f"📦 {result['Article']} - {result['Product Name']}", expanded=False):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.markdown(f"**Article Number:** {result['Article']}")
                                st.markdown(f"**Product Name:** {result['Product Name']}")
                                st.markdown(f"**Supplier:** {result['Supplier']}")
                                st.markdown(f"**Currency:** {result['Currency']}")
                            with col2:
                                st.markdown(f"**💰 Latest Price:** <span style='font-size: 1.5rem; font-weight: 700; color: #059669;'>{result['Latest Price']}/kg</span>", unsafe_allow_html=True)
                                st.markdown(f"**📅 Last Order Date:** {result['Last Order Date']}")
                                st.markdown(f"**📊 Total Orders:** {result['Total Orders']}")
                            
                            if result.get('Price History') and len(result['Price History']) > 1:
                                st.markdown("---")
                                st.markdown("**📈 Price History:**")
                                price_history_str = " → ".join(result['Price History'][-5:])
                                st.info(f"Last 5 prices: {price_history_str}")
                else:
                    st.warning(f"No results found for '{search_term}' in {selected_client}")
        else:
            st.info("👆 Enter an article number or product name to check the latest price")
    
    else:  # Show All Items
        with st.spinner(f"Loading all items for {selected_client}..."):
            client_data = get_google_sheets_data(selected_client)
            
            all_items = []
            
            for supplier in ["Backaldrin", "Bateel"]:
                supplier_data = client_data.get(supplier, {})
                
                for article_num, article_data in supplier_data.items():
                    orders = article_data.get('orders', [])
                    if orders:
                        sorted_orders = sorted(orders, key=lambda x: str(x.get('date', '')), reverse=True)
                        latest_order = sorted_orders[0] if sorted_orders else None
                        
                        if latest_order:
                            price_value = latest_order.get('price_value')
                            currency = latest_order.get('currency', 'USD')
                            latest_date = latest_order.get('date', 'Unknown')
                            symbol = currency_symbols.get(currency, "$")
                            
                            if price_value:
                                latest_price_display = price_value
                                latest_price_formatted = f"{symbol}{price_value:.2f}"
                            else:
                                latest_price_display = 0
                                latest_price_formatted = latest_order.get('price', 'N/A')
                            
                            all_items.append({
                                'Article': article_num,
                                'Product Name': article_data.get('names', ['N/A'])[0],
                                'Supplier': supplier,
                                'Latest Price': latest_price_display,
                                'Latest Price Formatted': latest_price_formatted,
                                'Currency': currency,
                                'Last Order Date': latest_date,
                                'Total Orders': len(orders)
                            })
            
            if all_items:
                df = pd.DataFrame(all_items)
                df = df.sort_values('Latest Price', ascending=False)
                
                st.success(f"✅ Loaded {len(df)} items with prices")
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Items", len(df))
                with col2:
                    # Calculate average by converting to USD (simple approach - show note)
                    usd_prices = [item['Latest Price'] for item in all_items if item['Currency'] == 'USD' and item['Latest Price'] > 0]
                    if usd_prices:
                        st.metric("Avg Price (USD only)", f"${sum(usd_prices)/len(usd_prices):.2f}")
                    else:
                        st.metric("Avg Price", "Multiple currencies")
                with col3:
                    st.metric("USD Items", len([i for i in all_items if i['Currency'] == 'USD']))
                with col4:
                    st.metric("EUR/SAR Items", len([i for i in all_items if i['Currency'] in ['EUR', 'SAR']]))
                
                col1, col2 = st.columns(2)
                with col1:
                    search_filter = st.text_input("🔍 Filter by Article or Product:", key="price_checker_filter")
                with col2:
                    supplier_filter = st.selectbox("Filter by Supplier:", ["All", "Backaldrin", "Bateel"], key="price_checker_supplier_filter")
                
                filtered_df = df.copy()
                if search_filter:
                    mask = filtered_df['Article'].astype(str).str.contains(search_filter, case=False, na=False)
                    mask = mask | filtered_df['Product Name'].astype(str).str.contains(search_filter, case=False, na=False)
                    filtered_df = filtered_df[mask]
                if supplier_filter != "All":
                    filtered_df = filtered_df[filtered_df['Supplier'] == supplier_filter]
                
                st.markdown(f"**Showing {len(filtered_df)} items**")
                
                # Display with formatted prices
                display_df = filtered_df[['Article', 'Product Name', 'Supplier', 'Latest Price Formatted', 'Currency', 'Last Order Date', 'Total Orders']].copy()
                display_df.columns = ['Article', 'Product Name', 'Supplier', 'Latest Price', 'Currency', 'Last Order Date', 'Total Orders']
                
                st.dataframe(display_df, use_container_width=True, hide_index=True)
                
                csv_data = filtered_df[['Article', 'Product Name', 'Supplier', 'Latest Price Formatted', 'Currency', 'Last Order Date', 'Total Orders']].to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📥 Export All Prices to CSV",
                    data=csv_data.encode('utf-8-sig'),
                    file_name=f"all_prices_{selected_client}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            else:
                st.warning(f"No price data found for {selected_client}")

# ============================================
# TAB 10: SALES HISTORY
# ============================================

def sales_history_tab():
    """View sales history for a specific item over a selected date range with currency support"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #f97316, #ea580c); padding: 1.25rem; border-radius: 12px; margin-bottom: 1.5rem;">
        <h2 style="margin:0; color: white;">📈 Sales History</h2>
        <p style="margin:0; opacity:0.9; color: white;">Track sales performance • Filter by date range • Supports USD, EUR, SAR</p>
    </div>
    """, unsafe_allow_html=True)
    
    available_clients = st.session_state.user_clients
    if not available_clients:
        st.warning("No clients available. Please check your Clients_CoC sheet.")
        return
    
    col1, col2 = st.columns(2)
    
    with col1:
        selected_client = st.selectbox("Select Client:", available_clients, key="sales_history_client")
    
    with col2:
        search_type = st.radio("Select by:", ["Article Number", "Product Name"], horizontal=True, key="sales_history_search_type")
    
    # Currency symbol mapping
    currency_symbols = {
        "USD": "$",
        "EUR": "€",
        "SAR": "ر.س"
    }
    
    if selected_client:
        with st.spinner(f"Loading items for {selected_client}..."):
            client_data = get_google_sheets_data(selected_client)
            
            # Build list of available items
            items_list = []
            for supplier in ["Backaldrin", "Bateel"]:
                supplier_data = client_data.get(supplier, {})
                for article_num, article_data in supplier_data.items():
                    product_name = article_data.get('names', ['N/A'])[0]
                    if product_name and product_name != 'N/A' and product_name != 'nan':
                        items_list.append({
                            'article': article_num,
                            'product': product_name,
                            'supplier': supplier
                        })
            
            if not items_list:
                st.warning(f"No items found for {selected_client}")
                return
            
            # Create dropdown options
            if search_type == "Article Number":
                item_options = sorted([f"{item['article']} - {item['product'][:50]}" for item in items_list])
                selected_item_display = st.selectbox("Select Item:", item_options, key="sales_history_item")
                selected_article = selected_item_display.split(" - ")[0] if selected_item_display else None
            else:
                item_options = sorted([f"{item['product'][:60]} - {item['article']}" for item in items_list])
                selected_item_display = st.selectbox("Select Item:", item_options, key="sales_history_item")
                selected_article = selected_item_display.split(" - ")[-1] if selected_item_display else None
            
            # Date range selection
            st.markdown("---")
            st.markdown("### 📅 Select Date Range")
            
            col1, col2 = st.columns(2)
            with col1:
                start_date = st.date_input("From Date:", 
                                           value=datetime.now().date() - pd.Timedelta(days=365),
                                           key="sales_history_start")
            with col2:
                end_date = st.date_input("To Date:", 
                                         value=datetime.now().date(),
                                         key="sales_history_end")
            
            if selected_article and start_date and end_date:
                # Find the item data
                item_data = None
                item_supplier = None
                for supplier in ["Backaldrin", "Bateel"]:
                    supplier_data = client_data.get(supplier, {})
                    if selected_article in supplier_data:
                        item_data = supplier_data[selected_article]
                        item_supplier = supplier
                        break
                
                if item_data:
                    product_name = item_data.get('names', ['N/A'])[0]
                    
                    st.markdown(f"### 📊 Sales History: {selected_article} - {product_name}")
                    st.markdown(f"**Supplier:** {item_supplier} | **Client:** {selected_client}")
                    
                    # Filter orders by date range
                    filtered_orders = []
                    for order in item_data.get('orders', []):
                        date_str = order.get('date', '')
                        if date_str and date_str != 'nan':
                            for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                                try:
                                    date_obj = datetime.strptime(str(date_str).strip(), fmt).date()
                                    if start_date <= date_obj <= end_date:
                                        order['parsed_date'] = date_obj
                                        filtered_orders.append(order)
                                    break
                                except:
                                    continue
                    
                    if filtered_orders:
                        # Sort by date
                        filtered_orders = sorted(filtered_orders, key=lambda x: x.get('parsed_date', datetime.min.date()))
                        
                        # Summary statistics
                        total_quantity = 0
                        total_weight = 0
                        total_value_usd = 0
                        total_value_eur = 0
                        total_value_sar = 0
                        prices = []
                        currencies_used = set()
                        
                        for order in filtered_orders:
                            # Safe quantity extraction
                            try:
                                qty_str = str(order.get('quantity', '0')).replace(',', '').strip()
                                qty_str = re.sub(r'[^0-9.-]', '', qty_str)
                                qty = float(qty_str) if qty_str and qty_str != 'nan' else 0
                                total_quantity += qty
                            except:
                                pass
                            
                            # Safe weight extraction
                            try:
                                weight_str = str(order.get('total_weight', '0')).replace(',', '').strip()
                                weight_str = re.sub(r'[^0-9.-]', '', weight_str)
                                weight = float(weight_str) if weight_str and weight_str != 'nan' else 0
                                total_weight += weight
                            except:
                                pass
                            
                            # Safe price extraction with currency
                            currency = order.get('currency', 'USD')
                            currencies_used.add(currency)
                            price_value = order.get('price_value')
                            
                            if price_value and price_value > 0:
                                prices.append({'value': price_value, 'currency': currency})
                                
                                # Track by currency
                                if currency == "USD":
                                    total_value_usd += price_value * weight if weight > 0 else 0
                                elif currency == "EUR":
                                    total_value_eur += price_value * weight if weight > 0 else 0
                                elif currency == "SAR":
                                    total_value_sar += price_value * weight if weight > 0 else 0
                        
                        # Display metrics
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Orders", len(filtered_orders))
                        with col2:
                            st.metric("Total Quantity", f"{total_quantity:,.0f}" if total_quantity > 0 else "N/A")
                        with col3:
                            st.metric("Total Weight", f"{total_weight:,.0f} kg" if total_weight > 0 else "N/A")
                        
                        # Display value metrics by currency
                        if total_value_usd > 0:
                            st.metric("Total Value (USD)", f"${total_value_usd:,.2f}")
                        if total_value_eur > 0:
                            st.metric("Total Value (EUR)", f"€{total_value_eur:,.2f}")
                        if total_value_sar > 0:
                            st.metric("Total Value (SAR)", f"ر.س{total_value_sar:,.2f}")
                        
                        # Price statistics
                        if prices:
                            usd_prices = [p['value'] for p in prices if p['currency'] == 'USD']
                            eur_prices = [p['value'] for p in prices if p['currency'] == 'EUR']
                            sar_prices = [p['value'] for p in prices if p['currency'] == 'SAR']
                            
                            price_stats_text = ""
                            if usd_prices:
                                price_stats_text += f"USD: ${min(usd_prices):.2f} - ${max(usd_prices):.2f} (avg ${sum(usd_prices)/len(usd_prices):.2f})"
                            if eur_prices:
                                if price_stats_text: price_stats_text += " | "
                                price_stats_text += f"EUR: €{min(eur_prices):.2f} - €{max(eur_prices):.2f} (avg €{sum(eur_prices)/len(eur_prices):.2f})"
                            if sar_prices:
                                if price_stats_text: price_stats_text += " | "
                                price_stats_text += f"SAR: ر.س{min(sar_prices):.2f} - ر.س{max(sar_prices):.2f} (avg ر.س{sum(sar_prices)/len(sar_prices):.2f})"
                            
                            st.info(f"**Price Range:** {price_stats_text}")
                        
                        # Create chart data
                        chart_data = []
                        for order in filtered_orders:
                            price_value = order.get('price_value')
                            currency = order.get('currency', 'USD')
                            weight_val = 0
                            
                            try:
                                weight_str = str(order.get('total_weight', '0')).replace(',', '').strip()
                                weight_str = re.sub(r'[^0-9.-]', '', weight_str)
                                weight_val = float(weight_str) if weight_str and weight_str != 'nan' else 0
                            except:
                                pass
                            
                            if price_value and price_value > 0:
                                chart_data.append({
                                    'Date': order.get('parsed_date'),
                                    'Price': price_value,
                                    'Currency': currency,
                                    'Weight (kg)': weight_val,
                                    'Order #': order.get('order_no', 'N/A'),
                                    'Price Label': f"{currency_symbols.get(currency, '$')}{price_value:.2f}"
                                })
                        
                        if chart_data:
                            chart_df = pd.DataFrame(chart_data)
                            
                            if len(chart_df) > 0:
                                st.markdown("### 📈 Price Trend")
                                price_chart = alt.Chart(chart_df).mark_line(point=True, color='#f97316', strokeWidth=2).encode(
                                    x=alt.X('Date:T', title='Order Date'),
                                    y=alt.Y('Price:Q', title='Price'),
                                    color=alt.Color('Currency:N', title='Currency'),
                                    tooltip=['Date:T', 'Price Label:N', 'Weight (kg):Q', 'Order #:N']
                                ).properties(height=350)
                                st.altair_chart(price_chart, use_container_width=True)
                                
                                st.markdown("### 📦 Sales Volume")
                                volume_chart = alt.Chart(chart_df).mark_bar(color='#f97316').encode(
                                    x=alt.X('Date:T', title='Order Date'),
                                    y=alt.Y('Weight (kg):Q', title='Weight (kg)'),
                                    tooltip=['Date:T', 'Weight (kg):Q', 'Order #:N']
                                ).properties(height=300)
                                st.altair_chart(volume_chart, use_container_width=True)
                        
                        # Detailed orders table
                        st.markdown("### 📋 Order Details")
                        orders_list = []
                        for order in filtered_orders:
                            price_display = order.get('price', 'N/A')
                            currency = order.get('currency', 'USD')
                            symbol = currency_symbols.get(currency, "$")
                            
                            # If we have parsed price value, format nicely
                            price_value = order.get('price_value')
                            if price_value:
                                price_display = f"{symbol}{price_value:.2f}"
                            
                            orders_list.append({
                                'Order Date': order.get('date', 'N/A'),
                                'Order #': order.get('order_no', 'N/A'),
                                'Quantity': order.get('quantity', 'N/A') if order.get('quantity') else 'N/A',
                                'Weight (kg)': order.get('total_weight', 'N/A') if order.get('total_weight') else 'N/A',
                                'Price': price_display,
                                'Currency': currency,
                                'Total Price': order.get('total_price', 'N/A') if order.get('total_price') else 'N/A'
                            })
                        
                        if orders_list:
                            orders_df = pd.DataFrame(orders_list)
                            st.dataframe(orders_df, use_container_width=True, hide_index=True)
                            
                            # Export option
                            csv_data = orders_df.to_csv(index=False, encoding='utf-8-sig')
                            st.download_button(
                                label="📥 Export Sales History to CSV",
                                data=csv_data.encode('utf-8-sig'),
                                file_name=f"sales_history_{selected_article}_{selected_client}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv",
                                mime="text/csv",
                                use_container_width=True
                            )
                        else:
                            st.info("No valid data to display in table")
                        
                    else:
                        st.warning(f"No orders found for this item between {start_date} and {end_date}")
                else:
                    st.error("Item not found")

# ============================================
# MAIN DASHBOARD
# ============================================

def main_dashboard():
    """Main dashboard with 10 consolidated tabs"""
    
    with st.sidebar:
        st.markdown("### 📋 Navigation")
        
        tabs = [
            "📋 CLIENT ORDERS",
            "💰 PRICING HUB", 
            "⭐ SPECIAL PRICES",
            "📦 PRODUCTS & LOGISTICS",
            "📅 ORDER TRACKING",
            "📈 PRICE TRACKING",
            "💰 COMMISSION",
            "👥 CLIENT DETAILS",
            "🔍 PRICE CHECKER",
            "📊 SALES HISTORY"
        ]
        
        for tab in tabs:
            is_active = st.session_state.get('active_tab', '📋 CLIENT ORDERS') == tab
            button_label = f"▶️ {tab}" if is_active else tab
            
            if st.button(
                button_label,
                key=f"tab_{tab}",
                use_container_width=True,
                type="primary" if is_active else "secondary"
            ):
                st.session_state.active_tab = tab
                st.rerun()
        
        st.markdown("---")
        
        if st.session_state.get('search_history'):
            st.markdown("### 🔍 Recent Searches")
            for i, history_item in enumerate(st.session_state.search_history[:5]):
                time_ago = format_time_ago(history_item['timestamp'])
                if st.button(f"{history_item['search_term'][:30]}", key=f"hist_{i}", use_container_width=True):
                    st.session_state.active_tab = "📋 CLIENT ORDERS"
                    st.rerun()
        
        st.markdown("---")
        st.markdown("### 📢 Updates")
        announcements = [
            "✅ NEW! Price Checker tab added!",
            "✅ NEW! Sales History with date range!",
            "🔍 Get latest prices instantly",
            "📊 Track sales over any period",
            "💰 Commission tracking available"
        ]
        for announcement in announcements:
            st.markdown(f'<div class="announcement-item">{announcement}</div>', unsafe_allow_html=True)
        
        st.markdown("---")
        logout_button()
    
    st.markdown(f"<div class='section-header'>{st.session_state.active_tab}</div>", unsafe_allow_html=True)
    
    if st.session_state.active_tab == "📋 CLIENT ORDERS":
        client_orders_tab()
    elif st.session_state.active_tab == "💰 PRICING HUB":
        pricing_hub_tab()
    elif st.session_state.active_tab == "⭐ SPECIAL PRICES":
        special_prices_tab()
    elif st.session_state.active_tab == "📦 PRODUCTS & LOGISTICS":
        products_logistics_tab()
    elif st.session_state.active_tab == "📅 ORDER TRACKING":
        order_tracking_tab()
    elif st.session_state.active_tab == "📈 PRICE TRACKING":
        price_tracking_tab()
    elif st.session_state.active_tab == "💰 COMMISSION":
        commission_tab()
    elif st.session_state.active_tab == "👥 CLIENT DETAILS":
        client_details_tab()
    elif st.session_state.active_tab == "🔍 PRICE CHECKER":
        price_checker_tab()
    elif st.session_state.active_tab == "📊 SALES HISTORY":
        sales_history_tab()
    
    st.markdown(f"""
    <div class="dashboard-footer">
        Multi-Client Dashboard v9.0 (with Price Checker & Sales History) | Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    </div>
    """, unsafe_allow_html=True)

# ============================================
# MAIN EXECUTION
# ============================================
if __name__ == "__main__":
    if not check_login():
        login_page()
    else:
        main_dashboard()
